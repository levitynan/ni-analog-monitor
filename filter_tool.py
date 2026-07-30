"""
filter_tool.py — Butterworth filter for main.py recordings

Load an .xlsx recording, apply a digital Butterworth filter, inspect the
frequency response, and save the filtered result.

Usage:
    python filter_tool.py [recording.xlsx]

Dependencies (in addition to requirements.txt):
    pip install scipy
"""

import pathlib
import sys
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from typing import Optional

import numpy as np
import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

try:
    from scipy.signal import butter, filtfilt, freqz
    HAS_SCIPY = True
    SCIPY_ERR  = ""
except Exception as _e:
    HAS_SCIPY = False
    SCIPY_ERR  = str(_e)

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

# ── Colours (Catppuccin Mocha) ────────────────────────────────────────────────
BG      = "#1e1e2e"
SURFACE = "#313244"
MUTED   = "#6c7086"
TEXT    = "#cdd6f4"
BLUE    = "#89b4fa"
GREEN   = "#a6e3a1"
YELLOW  = "#f9e2af"
MAUVE   = "#cba6f7"
BORDER  = "#45475a"

FILTER_TYPES = ["Low-pass", "High-pass", "Band-pass", "Band-stop"]
BTYPE_MAP    = {
    "Low-pass":   "lowpass",
    "High-pass":  "highpass",
    "Band-pass":  "bandpass",
    "Band-stop":  "bandstop",
}


class FilterApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Butterworth Filter — Recording Analyser")
        self.root.geometry("1060x900")
        self.root.configure(bg=BG)
        self.root.resizable(True, True)

        self._time:          Optional[np.ndarray] = None
        self._columns:       dict[str, np.ndarray] = {}
        self._fs:            float = 100.0
        self._filtered:      Optional[np.ndarray] = None
        self._plot_input:    Optional[np.ndarray] = None
        self._file_path      = ""
        self._apply_job:     Optional[str] = None
        self._region_spans:  list = []
        self._last_analysis: list = []

        self._build_ui()

        if len(sys.argv) > 1 and pathlib.Path(sys.argv[1]).exists():
            self._load_file(sys.argv[1])

    # ── UI construction ───────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        self._build_toolbar()
        self._build_params()
        self._build_analysis()
        self._build_statusbar()  # must be packed before the expanding plots frame
        self._build_plots()

    def _build_toolbar(self) -> None:
        tb = tk.Frame(self.root, bg=SURFACE)
        tb.pack(fill=tk.X, padx=20, pady=(12, 0))
        tk.Label(tb, text="File:", font=("Helvetica", 9),
                 fg=MUTED, bg=SURFACE).pack(side=tk.LEFT, padx=(8, 4), pady=6)
        self._path_var = tk.StringVar(value="No file loaded — click Browse or pass path as argument")
        tk.Label(tb, textvariable=self._path_var, font=("Helvetica", 9), fg=TEXT, bg=SURFACE,
                 anchor="w").pack(side=tk.LEFT, fill=tk.X, expand=True)
        tk.Button(tb, text="Browse…", command=self._browse,
                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=12, pady=4,
                  activebackground=MUTED, cursor="hand2").pack(side=tk.RIGHT, padx=8, pady=4)

    def _build_params(self) -> None:
        outer = tk.LabelFrame(self.root, text="  FILTER SETTINGS  ",
                              font=("Helvetica", 9, "bold"),
                              fg=BLUE, bg=SURFACE, bd=1, relief=tk.FLAT, labelanchor="nw")
        outer.pack(fill=tk.X, padx=20, pady=(10, 0))

        # ── Row 0: column, type, order, apply ──
        r0 = tk.Frame(outer, bg=SURFACE)
        r0.pack(fill=tk.X, padx=10, pady=(8, 4))

        tk.Label(r0, text="Column:", font=("Helvetica", 9), fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._col_var = tk.StringVar()
        self._col_combo = ttk.Combobox(r0, textvariable=self._col_var,
                                       state="readonly", width=22)
        self._col_combo.pack(side=tk.LEFT, padx=(4, 20))
        self._col_combo.bind("<<ComboboxSelected>>", self._on_col_change)

        tk.Label(r0, text="Type:", font=("Helvetica", 9), fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._type_var = tk.StringVar(value="Low-pass")
        type_cb = ttk.Combobox(r0, textvariable=self._type_var, state="readonly",
                               width=12, values=FILTER_TYPES)
        type_cb.pack(side=tk.LEFT, padx=(4, 20))
        type_cb.bind("<<ComboboxSelected>>", self._on_type_change)

        tk.Label(r0, text="Order:", font=("Helvetica", 9), fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._order_var = tk.IntVar(value=4)
        self._order_lbl = tk.Label(r0, text="4", font=("Courier New", 9, "bold"),
                                   fg=BLUE, bg=SURFACE, width=2)
        tk.Scale(r0, from_=1, to=8, resolution=1, orient=tk.HORIZONTAL,
                 variable=self._order_var, length=100, showvalue=False,
                 bg=SURFACE, fg=TEXT, troughcolor=BORDER, highlightthickness=0,
                 activebackground=BLUE,
                 command=lambda v: (self._order_lbl.configure(text=str(int(float(v)))),
                                    self._schedule_apply())
                 ).pack(side=tk.LEFT, padx=(4, 2))
        self._order_lbl.pack(side=tk.LEFT, padx=(0, 20))

        tk.Label(r0, text="Offset:", font=("Helvetica", 9), fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._offset_var = tk.StringVar(value="0.0")
        off_entry = tk.Entry(r0, textvariable=self._offset_var, width=8,
                             bg=BORDER, fg=TEXT, insertbackground=TEXT,
                             relief=tk.FLAT, font=("Courier New", 9))
        off_entry.pack(side=tk.LEFT, padx=(4, 2))
        off_entry.bind("<Return>", lambda _: self._apply_filter())
        off_entry.bind("<FocusOut>", lambda _: self._apply_filter())
        self._offset_unit_lbl = tk.Label(r0, text="", font=("Helvetica", 8),
                                         fg=MUTED, bg=SURFACE, width=18, anchor="w")
        self._offset_unit_lbl.pack(side=tk.LEFT, padx=(0, 10))

        self._auto_var = tk.BooleanVar(value=True)
        tk.Checkbutton(r0, text="Auto", variable=self._auto_var,
                       bg=SURFACE, fg=TEXT, selectcolor=BORDER,
                       activebackground=SURFACE, font=("Helvetica", 8)).pack(side=tk.RIGHT, padx=(0, 6))
        self._invert_var = tk.BooleanVar(value=False)
        tk.Checkbutton(r0, text="Invert", variable=self._invert_var,
                       command=self._apply_filter,
                       bg=SURFACE, fg=TEXT, selectcolor=BORDER,
                       activebackground=SURFACE, font=("Helvetica", 8)).pack(side=tk.RIGHT, padx=(0, 6))
        self._remove_dc_var = tk.BooleanVar(value=False)
        tk.Checkbutton(r0, text="Remove DC", variable=self._remove_dc_var,
                       command=self._apply_filter,
                       bg=SURFACE, fg=TEXT, selectcolor=BORDER,
                       activebackground=SURFACE, font=("Helvetica", 8)).pack(side=tk.RIGHT, padx=(0, 6))
        tk.Button(r0, text="Apply", command=self._apply_filter,
                  bg=BLUE, fg=BG, relief=tk.FLAT, padx=14, pady=3,
                  font=("Helvetica", 9, "bold"),
                  activebackground="#6ca0d0", cursor="hand2").pack(side=tk.RIGHT)

        # ── Row 1: fc1 (always visible) ──
        fc_outer = tk.Frame(outer, bg=SURFACE)
        fc_outer.pack(fill=tk.X, padx=10, pady=(0, 8))

        r1 = tk.Frame(fc_outer, bg=SURFACE)
        r1.pack(fill=tk.X)
        self._fc1_lbl_widget = tk.Label(r1, text="Cutoff:", font=("Helvetica", 9),
                                         fg=TEXT, bg=SURFACE, width=10, anchor="e")
        self._fc1_lbl_widget.pack(side=tk.LEFT)
        self._fc1_var = tk.DoubleVar(value=10.0)
        self._fc1_val_lbl = tk.Label(r1, text=" 10.00 Hz", font=("Courier New", 9),
                                      fg=BLUE, bg=SURFACE, width=10)
        self._fc1_scale = tk.Scale(
            r1, from_=0.01, to=50.0, resolution=0.01, orient=tk.HORIZONTAL,
            variable=self._fc1_var, length=440, showvalue=False,
            bg=SURFACE, fg=TEXT, troughcolor=BORDER, highlightthickness=0,
            activebackground=BLUE,
            command=lambda v: (self._fc1_val_lbl.configure(text=f"{float(v):6.2f} Hz"),
                               self._schedule_apply()))
        self._fc1_scale.pack(side=tk.LEFT, padx=(4, 2))
        self._fc1_val_lbl.pack(side=tk.LEFT, padx=(0, 12))
        self._nyq_lbl = tk.Label(r1, text="Nyquist: — Hz",
                                  font=("Helvetica", 8), fg=MUTED, bg=SURFACE)
        self._nyq_lbl.pack(side=tk.LEFT)

        # ── Row 2: fc2 (band-pass / band-stop only) ──
        self._fc2_frame = tk.Frame(fc_outer, bg=SURFACE)
        # not packed until needed
        tk.Label(self._fc2_frame, text="High cutoff:", font=("Helvetica", 9),
                 fg=TEXT, bg=SURFACE, width=10, anchor="e").pack(side=tk.LEFT)
        self._fc2_var = tk.DoubleVar(value=20.0)
        self._fc2_val_lbl = tk.Label(self._fc2_frame, text=" 20.00 Hz",
                                      font=("Courier New", 9), fg=BLUE, bg=SURFACE, width=10)
        self._fc2_scale = tk.Scale(
            self._fc2_frame, from_=0.01, to=50.0, resolution=0.01, orient=tk.HORIZONTAL,
            variable=self._fc2_var, length=440, showvalue=False,
            bg=SURFACE, fg=TEXT, troughcolor=BORDER, highlightthickness=0,
            activebackground=BLUE,
            command=lambda v: (self._fc2_val_lbl.configure(text=f"{float(v):6.2f} Hz"),
                               self._schedule_apply()))
        self._fc2_scale.pack(side=tk.LEFT, padx=(4, 2))
        self._fc2_val_lbl.pack(side=tk.LEFT)

    def _build_analysis(self) -> None:
        outer = tk.LabelFrame(self.root, text="  SIGNAL ANALYSIS  ",
                              font=("Helvetica", 9, "bold"),
                              fg=GREEN, bg=SURFACE, bd=1, relief=tk.FLAT, labelanchor="nw")
        outer.pack(fill=tk.X, padx=20, pady=(6, 0))

        ctrl = tk.Frame(outer, bg=SURFACE)
        ctrl.pack(fill=tk.X, padx=10, pady=(8, 4))

        tk.Label(ctrl, text="Window:", font=("Helvetica", 9),
                 fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._win_var = tk.DoubleVar(value=0.5)
        self._win_lbl = tk.Label(ctrl, text="0.50 s", font=("Courier New", 9),
                                  fg=GREEN, bg=SURFACE, width=7)
        tk.Scale(ctrl, from_=0.05, to=5.0, resolution=0.05, orient=tk.HORIZONTAL,
                 variable=self._win_var, length=140, showvalue=False,
                 bg=SURFACE, fg=TEXT, troughcolor=BORDER, highlightthickness=0,
                 activebackground=GREEN,
                 command=lambda v: self._win_lbl.configure(text=f"{float(v):.2f} s")
                 ).pack(side=tk.LEFT, padx=(4, 2))
        self._win_lbl.pack(side=tk.LEFT, padx=(0, 20))

        tk.Label(ctrl, text="Threshold:", font=("Helvetica", 9),
                 fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._thresh_var = tk.DoubleVar(value=0.2)
        self._thresh_lbl = tk.Label(ctrl, text="0.20", font=("Courier New", 9),
                                     fg=GREEN, bg=SURFACE, width=5)
        tk.Scale(ctrl, from_=0.01, to=1.0, resolution=0.01, orient=tk.HORIZONTAL,
                 variable=self._thresh_var, length=140, showvalue=False,
                 bg=SURFACE, fg=TEXT, troughcolor=BORDER, highlightthickness=0,
                 activebackground=GREEN,
                 command=lambda v: self._thresh_lbl.configure(text=f"{float(v):.2f}")
                 ).pack(side=tk.LEFT, padx=(4, 2))
        self._thresh_lbl.pack(side=tk.LEFT, padx=(0, 20))

        tk.Label(ctrl, text="Min value:", font=("Helvetica", 9),
                 fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._min_val_var = tk.StringVar(value="0.0")
        min_entry = tk.Entry(ctrl, textvariable=self._min_val_var, width=8,
                             bg=BORDER, fg=TEXT, insertbackground=TEXT,
                             relief=tk.FLAT, font=("Courier New", 9))
        min_entry.pack(side=tk.LEFT, padx=(4, 16))

        tk.Label(ctrl, text="Max regions:", font=("Helvetica", 9),
                 fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._max_regions_var = tk.StringVar(value="")
        tk.Entry(ctrl, textvariable=self._max_regions_var, width=5,
                 bg=BORDER, fg=TEXT, insertbackground=TEXT,
                 relief=tk.FLAT, font=("Courier New", 9)).pack(side=tk.LEFT, padx=(4, 4))
        tk.Label(ctrl, text="(blank = all)", font=("Helvetica", 8),
                 fg=MUTED, bg=SURFACE).pack(side=tk.LEFT, padx=(0, 4))

        tk.Button(ctrl, text="?", command=self._show_metrics_help,
                  bg=BORDER, fg=MAUVE, relief=tk.FLAT, padx=8, pady=3,
                  activebackground=MUTED, cursor="hand2",
                  font=("Helvetica", 9, "bold")).pack(side=tk.RIGHT, padx=(4, 0))
        tk.Button(ctrl, text="Clear", command=self._clear_analysis,
                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=10, pady=3,
                  activebackground=MUTED, cursor="hand2",
                  font=("Helvetica", 9)).pack(side=tk.RIGHT, padx=(4, 0))
        tk.Button(ctrl, text="Pub. export…", command=self._export_publication,
                  bg=MAUVE, fg=BG, relief=tk.FLAT, padx=10, pady=3,
                  activebackground=MUTED, cursor="hand2",
                  font=("Helvetica", 9, "bold")).pack(side=tk.RIGHT, padx=(4, 0))
        tk.Button(ctrl, text="Export image…", command=self._export_image,
                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=10, pady=3,
                  activebackground=MUTED, cursor="hand2",
                  font=("Helvetica", 9)).pack(side=tk.RIGHT, padx=(4, 0))
        tk.Button(ctrl, text="Analyse", command=self._analyse_regions,
                  bg=GREEN, fg=BG, relief=tk.FLAT, padx=14, pady=3,
                  font=("Helvetica", 9, "bold"),
                  activebackground="#7ec89d", cursor="hand2").pack(side=tk.RIGHT)

        self._analysis_text = tk.Text(outer, height=6, bg=BG, fg=TEXT,
                                       font=("Courier New", 9), relief=tk.FLAT,
                                       state=tk.DISABLED, wrap=tk.NONE)
        self._analysis_text.pack(fill=tk.X, padx=10, pady=(0, 8))
        self._analysis_text.tag_configure("header",  foreground=TEXT,   font=("Courier New", 9, "bold"))
        self._analysis_text.tag_configure("dynamic", foreground=YELLOW)
        self._analysis_text.tag_configure("static",  foreground=GREEN)
        self._analysis_text.tag_configure("muted",   foreground=MUTED)
        self._analysis_text.tag_configure("summary", foreground=MAUVE,  font=("Courier New", 9, "bold"))

    def _build_plots(self) -> None:
        pf = tk.Frame(self.root, bg=BG)
        pf.pack(fill=tk.BOTH, expand=True, padx=20, pady=(10, 0))

        # Signal plot
        sig_frame = tk.Frame(pf, bg=BG)
        sig_frame.pack(fill=tk.BOTH, expand=True)

        self._sig_fig, self._sig_ax = plt.subplots(figsize=(10, 3.2))
        self._sig_fig.patch.set_facecolor(BG)
        self._sig_ax.set_facecolor(SURFACE)
        self._sig_ax.tick_params(colors=TEXT, labelsize=8)
        for sp in self._sig_ax.spines.values():
            sp.set_color(BORDER)
        self._sig_ax.set_xlabel("Time (s)", color=TEXT, fontsize=9)
        self._sig_ax.set_ylabel("Value", color=TEXT, fontsize=9)
        self._sig_ax.grid(color=BORDER, linewidth=0.5, alpha=0.6)
        self._orig_line, = self._sig_ax.plot([], [], color=MUTED, linewidth=1.0,
                                               alpha=0.55, label="Original")
        self._filt_line, = self._sig_ax.plot([], [], color=BLUE, linewidth=1.6,
                                               label="Filtered")
        self._sig_ax.legend(facecolor=SURFACE, edgecolor=BORDER, labelcolor=TEXT, fontsize=8)
        self._sig_fig.tight_layout(pad=1.0)
        self._sig_canvas = FigureCanvasTkAgg(self._sig_fig, master=sig_frame)
        self._sig_toolbar = NavigationToolbar2Tk(self._sig_canvas, sig_frame)
        self._sig_toolbar.update()
        self._style_toolbar(self._sig_toolbar)
        self._sig_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

        # Frequency response (Bode magnitude) plot
        bode_frame = tk.Frame(pf, bg=BG)
        bode_frame.pack(fill=tk.BOTH, expand=True)

        self._bode_fig, self._bode_ax = plt.subplots(figsize=(10, 2.4))
        self._bode_fig.patch.set_facecolor(BG)
        self._bode_ax.set_facecolor(SURFACE)
        self._bode_ax.tick_params(colors=TEXT, labelsize=8)
        for sp in self._bode_ax.spines.values():
            sp.set_color(BORDER)
        self._bode_ax.set_xlabel("Frequency (Hz)", color=TEXT, fontsize=9)
        self._bode_ax.set_ylabel("Magnitude (dB)", color=TEXT, fontsize=9)
        self._bode_ax.grid(color=BORDER, linewidth=0.5, alpha=0.6)
        self._bode_ax.axhline(-3, color=YELLOW, linewidth=0.8,
                               linestyle="--", alpha=0.7, label="−3 dB")
        self._bode_line, = self._bode_ax.plot([], [], color=MAUVE, linewidth=1.5)
        self._bode_ax.legend(facecolor=SURFACE, edgecolor=BORDER, labelcolor=TEXT, fontsize=8)
        self._bode_fig.tight_layout(pad=1.0)
        self._bode_canvas = FigureCanvasTkAgg(self._bode_fig, master=bode_frame)
        self._bode_toolbar = NavigationToolbar2Tk(self._bode_canvas, bode_frame)
        self._bode_toolbar.update()
        self._style_toolbar(self._bode_toolbar)
        self._bode_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def _style_toolbar(self, toolbar: NavigationToolbar2Tk) -> None:
        toolbar.configure(background=SURFACE)
        for child in toolbar.winfo_children():
            try:
                child.configure(background=SURFACE, foreground=TEXT,
                                highlightbackground=SURFACE, relief=tk.FLAT)
            except tk.TclError:
                pass

    # ── Signal analysis ───────────────────────────────────────────────────────

    def _analyse_regions(self) -> None:
        signal = self._filtered
        if signal is None or self._time is None:
            messagebox.showinfo("No data", "Apply a filter first.")
            return

        from scipy.ndimage import uniform_filter1d

        time = self._time
        win_sec = self._win_var.get()
        threshold_factor = self._thresh_var.get()
        window = max(3, int(win_sec * self._fs))

        # Activity = smoothed absolute derivative (physical rate of change per second)
        dt = 1.0 / self._fs
        derivative = np.abs(np.gradient(signal, dt))
        activity = uniform_filter1d(derivative, size=window)

        act_max = activity.max()
        if act_max < 1e-12:
            messagebox.showinfo("Flat signal", "Signal is constant — no dynamic regions found.")
            return
        activity_norm = activity / act_max

        is_dynamic = activity_norm > threshold_factor

        # Find contiguous regions
        raw_regions: list[tuple[int, int, str]] = []
        i, n = 0, len(is_dynamic)
        while i < n:
            rtype = "dynamic" if is_dynamic[i] else "quasi-static"
            j = i + 1
            while j < n and is_dynamic[j] == is_dynamic[i]:
                j += 1
            raw_regions.append((i, j - 1, rtype))
            i = j

        # Merge regions shorter than 10 % of the window into their neighbour
        min_samples = max(3, int(0.1 * window))
        regions: list[tuple[int, int, str]] = []
        for r in raw_regions:
            s, e, rt = r
            if (e - s + 1) < min_samples and regions:
                ps, _, pt = regions[-1]
                regions[-1] = (ps, e, pt)
            else:
                regions.append(r)

        # Compute per-region statistics
        try:
            min_val = float(self._min_val_var.get())
        except ValueError:
            min_val = 0.0

        results = []
        for s, e, rt in regions:
            seg = signal[s:e + 1]
            peak = float(seg.max())
            if peak < min_val:
                continue
            results.append({
                "type":     rt,
                "t_start":  float(time[s]),
                "t_end":    float(time[e]),
                "duration": float(time[e] - time[s]),
                "max":      peak,
                "min":      float(seg.min()),
                "mean":     float(seg.mean()),
                "std":      float(seg.std()),
                "peak_pk":  float(seg.max() - seg.min()),
            })

        # Limit to N regions with highest peak value, re-sorted by time
        try:
            max_n = int(self._max_regions_var.get())
            if max_n > 0:
                results = sorted(results, key=lambda r: r["max"], reverse=True)[:max_n]
                results = sorted(results, key=lambda r: r["t_start"])
        except ValueError:
            pass  # blank or invalid → keep all

        # Redraw only kept regions (spans were drawn before filtering)
        self._clear_analysis_spans()
        colour_map2 = {"dynamic": YELLOW, "quasi-static": GREEN}
        for r in results:
            span = self._sig_ax.axvspan(
                r["t_start"], r["t_end"], alpha=0.13,
                color=colour_map2[r["type"]], linewidth=0, zorder=0)
            self._region_spans.append(span)
        self._sig_canvas.draw_idle()

        self._last_analysis = results
        self._show_analysis_results(results)

    def _show_analysis_results(self, results: list) -> None:
        col = self._col_var.get()
        hdr = (f"{'#':>3}  {'Type':<14}  {'t-start':>8}  {'t-end':>8}  "
               f"{'Dur(s)':>7}  {'Max':>10}  {'Min':>10}  "
               f"{'Mean':>10}  {'Std':>10}  {'Pk-Pk':>10}")
        sep = "─" * len(hdr)

        self._analysis_text.configure(state=tk.NORMAL)
        self._analysis_text.delete("1.0", tk.END)
        self._analysis_text.insert(tk.END, f"Column: {col}\n", "muted")
        self._analysis_text.insert(tk.END, hdr + "\n", "header")
        self._analysis_text.insert(tk.END, sep + "\n", "muted")

        dyn_count = qs_count = 0
        dyn_dur   = qs_dur   = 0.0

        for i, r in enumerate(results, start=1):
            tag  = "dynamic" if r["type"] == "dynamic" else "static"
            label = "Dynamic" if r["type"] == "dynamic" else "Quasi-static"
            line = (f"{i:>3}  {label:<14}  {r['t_start']:>8.3f}  {r['t_end']:>8.3f}  "
                    f"{r['duration']:>7.3f}  {r['max']:>10.4f}  {r['min']:>10.4f}  "
                    f"{r['mean']:>10.4f}  {r['std']:>10.4f}  {r['peak_pk']:>10.4f}\n")
            self._analysis_text.insert(tk.END, line, tag)
            if r["type"] == "dynamic":
                dyn_count += 1; dyn_dur += r["duration"]
            else:
                qs_count += 1;  qs_dur  += r["duration"]

        self._analysis_text.insert(tk.END, sep + "\n", "muted")
        summary = (f"Dynamic: {dyn_count} region(s), {dyn_dur:.3f} s total  |  "
                   f"Quasi-static: {qs_count} region(s), {qs_dur:.3f} s total\n")
        self._analysis_text.insert(tk.END, summary, "summary")
        self._analysis_text.configure(state=tk.DISABLED)

    def _clear_analysis_spans(self) -> None:
        for span in self._region_spans:
            try:
                span.remove()
            except Exception:
                pass
        self._region_spans.clear()
        if hasattr(self, "_sig_canvas"):
            self._sig_canvas.draw_idle()

    def _clear_analysis(self) -> None:
        self._clear_analysis_spans()
        self._analysis_text.configure(state=tk.NORMAL)
        self._analysis_text.delete("1.0", tk.END)
        self._analysis_text.configure(state=tk.DISABLED)

    def _export_image(self) -> None:
        if self._filtered is None or self._time is None:
            messagebox.showinfo("No data", "Apply a filter first.")
            return

        stem = pathlib.Path(self._file_path).stem if self._file_path else "signal"
        default = f"{stem}_analysis.png"
        path = filedialog.asksaveasfilename(
            title="Export signal image",
            defaultextension=".png",
            initialfile=default,
            filetypes=[
                ("PNG image",      "*.png"),
                ("PDF document",   "*.pdf"),
                ("SVG vector",     "*.svg"),
                ("All files",      "*.*"),
            ],
        )
        if not path:
            return

        col_name     = self._col_var.get()
        has_analysis = bool(self._last_analysis)

        # ── Build figure ─────────────────────────────────────────────────────
        if has_analysis:
            fig = plt.figure(figsize=(14, 10), facecolor=BG)
            gs  = fig.add_gridspec(2, 1, height_ratios=[3, 1], hspace=0.45)
            ax_sig = fig.add_subplot(gs[0])
            ax_tbl = fig.add_subplot(gs[1])
        else:
            fig    = plt.figure(figsize=(14, 6), facecolor=BG)
            ax_sig = fig.add_subplot(111)
            ax_tbl = None

        # ── Signal axes ───────────────────────────────────────────────────────
        ax_sig.set_facecolor(SURFACE)
        ax_sig.tick_params(colors=TEXT, labelsize=9)
        for sp in ax_sig.spines.values():
            sp.set_color(BORDER)
        ax_sig.set_xlabel("Time (s)", color=TEXT, fontsize=10)
        ax_sig.set_ylabel(col_name,   color=TEXT, fontsize=10)
        ax_sig.grid(color=BORDER, linewidth=0.5, alpha=0.6)

        if self._plot_input is not None:
            ax_sig.plot(self._time, self._plot_input,
                        color=MUTED, linewidth=1.0, alpha=0.55, label="Original")
        ax_sig.plot(self._time, self._filtered,
                    color=BLUE, linewidth=1.6, label="Filtered")

        colour_map = {"dynamic": YELLOW, "quasi-static": GREEN}
        for r in self._last_analysis:
            ax_sig.axvspan(r["t_start"], r["t_end"],
                           alpha=0.15, color=colour_map[r["type"]],
                           linewidth=0, zorder=0)
            # Label each region at the top (axes-fraction y=1) with type + mean
            mid_t = (r["t_start"] + r["t_end"]) / 2.0
            label = ("Dyn" if r["type"] == "dynamic" else "QS") + f"\nμ={r['mean']:.3f}"
            ax_sig.text(mid_t, 0.97, label,
                        color=colour_map[r["type"]], fontsize=7,
                        ha="center", va="top",
                        transform=ax_sig.get_xaxis_transform())

        # Filter settings in title
        btype = self._type_var.get()
        order = self._order_var.get()
        fc1   = self._fc1_var.get()
        if btype in ("Band-pass", "Band-stop"):
            fc_str = f"{fc1:.2f} – {self._fc2_var.get():.2f} Hz"
        else:
            fc_str = f"{fc1:.2f} Hz"
        try:
            offset = float(self._offset_var.get())
        except ValueError:
            offset = 0.0
        offset_str = f"  |  offset {offset:+.4g} {col_name}" if offset != 0.0 else ""
        title = (f"{pathlib.Path(self._file_path).name}   ·   {col_name}\n"
                 f"{btype}  |  order {order}  |  fc = {fc_str}  |  fs ≈ {self._fs:.1f} Hz{offset_str}")
        ax_sig.set_title(title, color=TEXT, fontsize=10, pad=8)
        ax_sig.legend(facecolor=SURFACE, edgecolor=BORDER, labelcolor=TEXT, fontsize=9)
        ax_sig.autoscale_view()

        # ── Analysis table ────────────────────────────────────────────────────
        if ax_tbl is not None and has_analysis:
            ax_tbl.set_facecolor(BG)
            ax_tbl.axis("off")

            col_labels = ["#", "Type", "t-start (s)", "t-end (s)",
                          "Duration (s)", "Max", "Min", "Mean", "Std Dev", "Pk-Pk"]
            cell_data  = []
            row_colours: list[list[str]] = []
            for i, r in enumerate(self._last_analysis, start=1):
                label = "Dynamic" if r["type"] == "dynamic" else "Quasi-static"
                cell_data.append([
                    str(i), label,
                    f"{r['t_start']:.3f}",  f"{r['t_end']:.3f}",
                    f"{r['duration']:.3f}",
                    f"{r['max']:.4f}",       f"{r['min']:.4f}",
                    f"{r['mean']:.4f}",      f"{r['std']:.4f}",
                    f"{r['peak_pk']:.4f}",
                ])
                c = "#3a3020" if r["type"] == "dynamic" else "#1e2e22"
                row_colours.append([c] * len(col_labels))

            tbl = ax_tbl.table(
                cellText=cell_data,
                colLabels=col_labels,
                cellLoc="center",
                loc="center",
            )
            tbl.auto_set_font_size(False)
            tbl.set_fontsize(8.5)
            tbl.scale(1, 1.6)

            for (ri, ci), cell in tbl.get_celld().items():
                cell.set_edgecolor(BORDER)
                if ri == 0:
                    cell.set_facecolor(BORDER)
                    cell.set_text_props(color=TEXT, fontweight="bold")
                else:
                    cell.set_facecolor(row_colours[ri - 1][ci])
                    r_data = self._last_analysis[ri - 1]
                    cell.set_text_props(
                        color=YELLOW if r_data["type"] == "dynamic" else GREEN)

            dyn = [r for r in self._last_analysis if r["type"] == "dynamic"]
            qs  = [r for r in self._last_analysis if r["type"] == "quasi-static"]
            summary = (f"Dynamic: {len(dyn)} region(s), "
                       f"{sum(r['duration'] for r in dyn):.3f} s total   |   "
                       f"Quasi-static: {len(qs)} region(s), "
                       f"{sum(r['duration'] for r in qs):.3f} s total")
            ax_tbl.text(0.5, -0.08, summary, transform=ax_tbl.transAxes,
                        color=MAUVE, fontsize=9, ha="center", va="top",
                        fontweight="bold")

        fig.tight_layout()
        try:
            fig.savefig(path, dpi=150, bbox_inches="tight", facecolor=BG)
            self._status_var.set(f"Image exported  →  {pathlib.Path(path).name}")
        except Exception as exc:
            messagebox.showerror("Export failed", str(exc))
        finally:
            plt.close(fig)

    def _export_publication(self) -> None:
        """Open publication export options, then call _do_pub_export."""
        if self._filtered is None or self._time is None:
            messagebox.showinfo("No data", "Apply a filter first.")
            return

        dlg = tk.Toplevel(self.root)
        dlg.title("Publication Export")
        dlg.configure(bg=BG)
        dlg.geometry("430x310")
        dlg.resizable(False, False)
        dlg.transient(self.root)
        dlg.grab_set()

        tk.Label(dlg, text="Publication Export Settings",
                 font=("Helvetica", 12, "bold"), fg=TEXT, bg=BG).pack(pady=(14, 2))
        tk.Label(dlg, text="White background · print-ready PNG",
                 font=("Helvetica", 8), fg=MUTED, bg=BG).pack(pady=(0, 8))

        grid = tk.Frame(dlg, bg=SURFACE)
        grid.pack(fill=tk.X, padx=20)

        def _row(label, row):
            tk.Label(grid, text=label, font=("Helvetica", 9), fg=TEXT, bg=SURFACE,
                     anchor="w", width=14).grid(row=row, column=0,
                                                sticky="w", padx=(10, 4), pady=5)

        def _rbframe(row):
            f = tk.Frame(grid, bg=SURFACE)
            f.grid(row=row, column=1, sticky="w", pady=5, padx=(0, 10))
            return f

        def _rb(parent, text, var, value):
            tk.Radiobutton(parent, text=text, variable=var, value=value,
                           bg=SURFACE, fg=TEXT, selectcolor=BORDER,
                           activebackground=SURFACE,
                           font=("Helvetica", 8)).pack(side=tk.LEFT, padx=(0, 8))

        # Figure width
        _row("Figure width:", 0)
        width_var = tk.StringVar(value="double")
        wf = _rbframe(0)
        _rb(wf, 'Single col (3.5")', width_var, "single")
        _rb(wf, 'Double col (7.25")', width_var, "double")
        _rb(wf, "Custom:", width_var, "custom")
        custom_w_var = tk.StringVar(value="5.0")
        tk.Entry(wf, textvariable=custom_w_var, width=5,
                 bg=BORDER, fg=TEXT, insertbackground=TEXT,
                 relief=tk.FLAT, font=("Courier New", 9)).pack(side=tk.LEFT)
        tk.Label(wf, text=" in", font=("Helvetica", 8),
                 fg=MUTED, bg=SURFACE).pack(side=tk.LEFT)

        # DPI
        _row("Resolution:", 1)
        dpi_var = tk.StringVar(value="300")
        df = _rbframe(1)
        for v in ("300", "600"):
            _rb(df, f"{v} dpi", dpi_var, v)

        # Colour mode
        _row("Colour mode:", 2)
        color_var = tk.StringVar(value="color")
        cf = _rbframe(2)
        _rb(cf, "Colour",     color_var, "color")
        _rb(cf, "Greyscale",  color_var, "grey")

        # Font size
        _row("Font size:", 3)
        font_var = tk.StringVar(value="8")
        ff = _rbframe(3)
        for v in ("7", "8", "9", "10"):
            _rb(ff, f"{v} pt", font_var, v)

        # Analysis table checkbox (only shown when analysis exists)
        table_var = tk.BooleanVar(value=bool(self._last_analysis))
        if self._last_analysis:
            tk.Checkbutton(dlg, text="Include analysis table",
                           variable=table_var,
                           bg=BG, fg=TEXT, selectcolor=BORDER,
                           activebackground=BG,
                           font=("Helvetica", 8)).pack(anchor="w", padx=24, pady=(6, 0))

        def _do() -> None:
            wmap = {"single": 3.5, "double": 7.25}
            sel  = width_var.get()
            if sel in wmap:
                fig_w = wmap[sel]
            else:
                try:
                    fig_w = float(custom_w_var.get())
                    if fig_w <= 0:
                        raise ValueError
                except ValueError:
                    messagebox.showerror("Invalid width",
                                         "Enter a positive number for width.",
                                         parent=dlg)
                    return
            opts = {
                "width":         fig_w,
                "dpi":           int(dpi_var.get()),
                "color_mode":    color_var.get(),
                "font_size":     int(font_var.get()),
                "include_table": table_var.get() if self._last_analysis else False,
            }
            dlg.destroy()
            self._do_pub_export(opts)

        btns = tk.Frame(dlg, bg=BG)
        btns.pack(fill=tk.X, padx=20, pady=(10, 16))
        tk.Button(btns, text="Export PNG…", command=_do,
                  bg=MAUVE, fg=BG, relief=tk.FLAT, padx=14, pady=4,
                  activebackground=MUTED, cursor="hand2",
                  font=("Helvetica", 9, "bold")).pack(side=tk.LEFT)
        tk.Button(btns, text="Cancel", command=dlg.destroy,
                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=12, pady=4,
                  activebackground=MUTED, cursor="hand2").pack(side=tk.LEFT, padx=(8, 0))

    def _do_pub_export(self, opts: dict) -> None:
        """Build and save a publication-quality white-background figure."""
        import matplotlib as mpl

        fig_width     = opts["width"]
        dpi           = opts["dpi"]
        greyscale     = opts["color_mode"] == "grey"
        font_size     = opts["font_size"]
        has_table     = opts["include_table"] and bool(self._last_analysis)
        col_name      = self._col_var.get()

        stem    = pathlib.Path(self._file_path).stem if self._file_path else "signal"
        default = f"{stem}_pub.png"
        path    = filedialog.asksaveasfilename(
            title="Save publication figure",
            defaultextension=".png",
            initialfile=default,
            filetypes=[("PNG image", "*.png"), ("All files", "*.*")],
        )
        if not path:
            return

        # Colour palette (colour vs greyscale)
        if greyscale:
            c_orig    = "#AAAAAA"
            c_filt    = "#000000"
            c_dyn_bg  = "#DDDDDD"
            c_qs_bg   = None       # no fill; white background distinguishes from hatched
            c_dyn_lbl = "#444444"
            c_qs_lbl  = "#444444"
            c_dyn_tbl = "#DDDDDD"
            c_qs_tbl  = "#F5F5F5"
            c_hdr_tbl = "#333333"
        else:
            c_orig    = "#AAAAAA"
            c_filt    = "#1565C0"   # dark blue
            c_dyn_bg  = "#FFF3CD"   # light amber
            c_qs_bg   = "#E8F5E9"   # light green
            c_dyn_lbl = "#B8860B"   # dark goldenrod
            c_qs_lbl  = "#2E7D32"   # dark green
            c_dyn_tbl = "#FFF3CD"
            c_qs_tbl  = "#E8F5E9"
            c_hdr_tbl = "#1565C0"

        # Figure height: signal panel is 0.6 × width; table appended below
        sig_h = fig_width * 0.60
        if has_table:
            tbl_h = max(1.2, len(self._last_analysis) * 0.22 + 0.6)
            fig_h = sig_h + tbl_h + 0.4
        else:
            fig_h = sig_h

        rc = {
            "font.size":            font_size,
            "axes.labelsize":       font_size,
            "xtick.labelsize":      font_size - 1,
            "ytick.labelsize":      font_size - 1,
            "legend.fontsize":      font_size - 1,
            "axes.titlesize":       font_size,
            "font.family":          "sans-serif",
            "axes.linewidth":       0.7,
            "grid.linewidth":       0.4,
            "grid.color":           "#CCCCCC",
            "grid.alpha":           1.0,
            "axes.grid":            True,
            "lines.linewidth":      1.2,
            "xtick.direction":      "in",
            "ytick.direction":      "in",
            "xtick.major.size":     3.0,
            "ytick.major.size":     3.0,
            "legend.framealpha":    1.0,
            "legend.edgecolor":     "#CCCCCC",
            "figure.facecolor":     "white",
            "axes.facecolor":       "white",
            "text.color":           "black",
            "axes.labelcolor":      "black",
            "xtick.color":          "black",
            "ytick.color":          "black",
        }

        with mpl.rc_context(rc):
            if has_table:
                fig = plt.figure(figsize=(fig_width, fig_h), facecolor="white")
                gs  = fig.add_gridspec(2, 1,
                                       height_ratios=[sig_h, tbl_h],
                                       hspace=0.5)
                ax_sig = fig.add_subplot(gs[0])
                ax_tbl = fig.add_subplot(gs[1])
            else:
                fig    = plt.figure(figsize=(fig_width, fig_h), facecolor="white")
                ax_sig = fig.add_subplot(111)
                ax_tbl = None

            # ── Signal panel ─────────────────────────────────────────────
            ax_sig.set_facecolor("white")
            ax_sig.spines["top"].set_visible(False)
            ax_sig.spines["right"].set_visible(False)
            for side in ("bottom", "left"):
                ax_sig.spines[side].set_linewidth(0.7)
                ax_sig.spines[side].set_color("black")
            ax_sig.set_xlabel("Time (s)", color="black")
            ax_sig.set_ylabel(col_name, color="black")

            if self._plot_input is not None:
                ax_sig.plot(self._time, self._plot_input,
                            color=c_orig, linewidth=0.8,
                            linestyle="--" if greyscale else "-",
                            alpha=0.7, label="Original")
            ax_sig.plot(self._time, self._filtered,
                        color=c_filt, linewidth=1.2, label="Filtered")

            # Region shading and labels
            for r in self._last_analysis:
                is_dyn = r["type"] == "dynamic"
                ax_sig.axvspan(
                    r["t_start"], r["t_end"],
                    facecolor=c_dyn_bg if is_dyn else (c_qs_bg or "none"),
                    hatch="////" if (greyscale and is_dyn) else "",
                    edgecolor="#888888" if (greyscale and is_dyn) else "none",
                    alpha=0.55, linewidth=0.0, zorder=0)
                mid_t = (r["t_start"] + r["t_end"]) / 2.0
                lbl   = ("D" if is_dyn else "QS") + f"\nμ={r['mean']:.3f}"
                ax_sig.text(mid_t, 0.97, lbl,
                            color=c_dyn_lbl if is_dyn else c_qs_lbl,
                            fontsize=max(5, font_size - 2),
                            ha="center", va="top",
                            transform=ax_sig.get_xaxis_transform())

            # Title with filter settings
            btype  = self._type_var.get()
            order  = self._order_var.get()
            fc1    = self._fc1_var.get()
            fc_str = (f"{fc1:.2f} – {self._fc2_var.get():.2f} Hz"
                      if btype in ("Band-pass", "Band-stop") else f"{fc1:.2f} Hz")
            try:
                offset = float(self._offset_var.get())
            except ValueError:
                offset = 0.0
            offset_str = f"  |  offset {offset:+.4g} {col_name}" if offset != 0.0 else ""
            ax_sig.set_title(
                f"{pathlib.Path(self._file_path).name}   ·   {col_name}\n"
                f"{btype}  |  order {order}  |  fc = {fc_str}"
                f"  |  fs ≈ {self._fs:.1f} Hz{offset_str}",
                color="black", pad=6)
            ax_sig.legend(framealpha=1.0, edgecolor="#CCCCCC")
            ax_sig.autoscale_view()

            # ── Analysis summary line ─────────────────────────────────────
            if self._last_analysis:
                dyn_r = [r for r in self._last_analysis if r["type"] == "dynamic"]
                qs_r  = [r for r in self._last_analysis if r["type"] == "quasi-static"]
                sig_mean = float(np.mean(self._filtered))
                sig_std  = float(np.std(self._filtered))
                sig_ppk  = float(np.max(self._filtered) - np.min(self._filtered))
                summary = (
                    f"Dynamic: {len(dyn_r)} region(s), "
                    f"{sum(r['duration'] for r in dyn_r):.3f} s"
                    f"   ·   "
                    f"Quasi-static: {len(qs_r)} region(s), "
                    f"{sum(r['duration'] for r in qs_r):.3f} s"
                    f"   ·   "
                    f"Signal  mean={sig_mean:.4f}  std={sig_std:.4f}  pk-pk={sig_ppk:.4f} {col_name}"
                )
                ax_sig.annotate(
                    summary,
                    xy=(0.5, -0.22), xycoords="axes fraction",
                    ha="center", va="top",
                    fontsize=max(5, font_size - 2),
                    color="#444444", style="italic")

            # ── Analysis table panel ──────────────────────────────────────
            if ax_tbl is not None and has_table:
                ax_tbl.set_facecolor("white")
                ax_tbl.axis("off")
                col_labels = ["#", "Type", "t-start (s)", "t-end (s)",
                              "Dur (s)", "Max", "Min", "Mean", "Std Dev", "Pk-Pk"]
                cell_data: list[list[str]] = []
                row_cols:  list[list[str]] = []
                for i, r in enumerate(self._last_analysis, start=1):
                    label = "Dynamic" if r["type"] == "dynamic" else "Quasi-static"
                    cell_data.append([
                        str(i), label,
                        f"{r['t_start']:.3f}", f"{r['t_end']:.3f}",
                        f"{r['duration']:.3f}",
                        f"{r['max']:.4f}", f"{r['min']:.4f}",
                        f"{r['mean']:.4f}", f"{r['std']:.4f}",
                        f"{r['peak_pk']:.4f}",
                    ])
                    c = c_dyn_tbl if r["type"] == "dynamic" else c_qs_tbl
                    row_cols.append([c] * len(col_labels))

                tbl = ax_tbl.table(cellText=cell_data, colLabels=col_labels,
                                   cellLoc="center", loc="center")
                tbl.auto_set_font_size(False)
                tbl.set_fontsize(font_size - 1)
                tbl.scale(1, 1.5)
                for (ri, ci), cell in tbl.get_celld().items():
                    cell.set_edgecolor("#AAAAAA")
                    cell.set_linewidth(0.5)
                    if ri == 0:
                        cell.set_facecolor(c_hdr_tbl)
                        cell.set_text_props(color="white", fontweight="bold")
                    else:
                        cell.set_facecolor(row_cols[ri - 1][ci])
                        cell.set_text_props(color="black")

                dyn = [r for r in self._last_analysis if r["type"] == "dynamic"]
                qs  = [r for r in self._last_analysis if r["type"] == "quasi-static"]
                ax_tbl.text(
                    0.5, -0.05,
                    f"Dynamic: {len(dyn)} region(s), "
                    f"{sum(r['duration'] for r in dyn):.3f} s   |   "
                    f"Quasi-static: {len(qs)} region(s), "
                    f"{sum(r['duration'] for r in qs):.3f} s",
                    transform=ax_tbl.transAxes, color="black",
                    fontsize=font_size - 1, ha="center", va="top")

            fig.tight_layout()
            try:
                fig.savefig(path, dpi=dpi, bbox_inches="tight", facecolor="white")
                self._status_var.set(
                    f"Publication figure saved  →  {pathlib.Path(path).name}"
                    f"  ({fig_width:.2f}\" × {fig_h:.2f}\"  @  {dpi} dpi)")
            except Exception as exc:
                messagebox.showerror("Export failed", str(exc))
            finally:
                plt.close(fig)

    def _show_metrics_help(self) -> None:
        win = tk.Toplevel(self.root)
        win.title("Analysis metrics — help")
        win.configure(bg=BG)
        win.geometry("620x560")
        win.resizable(False, False)
        win.transient(self.root)
        win.grab_set()

        tk.Label(win, text="Signal Analysis — Metrics Guide",
                 font=("Helvetica", 13, "bold"), fg=TEXT, bg=BG).pack(pady=(16, 2))
        tk.Label(win, text="Definitions for every column shown in the results table.",
                 font=("Helvetica", 9), fg=MUTED, bg=BG).pack(pady=(0, 10))

        sections = [
            ("Region classification", MAUVE, [
                ("Dynamic",
                 "A region where the signal is changing rapidly. Identified by the smoothed "
                 "absolute derivative exceeding the Threshold fraction of its peak value. "
                 "Shown with yellow shading."),
                ("Quasi-static",
                 "A region where the signal is settling or holding steady (low rate of change). "
                 "Shown with green shading."),
            ]),
            ("Time metrics", BLUE, [
                ("t-start (s)",  "Time stamp at the start of the region."),
                ("t-end (s)",    "Time stamp at the end of the region."),
                ("Duration (s)", "Length of the region in seconds (t-end − t-start)."),
            ]),
            ("Amplitude metrics", GREEN, [
                ("Max",
                 "Highest signal value within the region. Used by the Min value filter — "
                 "regions whose Max is below Min value are excluded."),
                ("Min",   "Lowest signal value within the region."),
                ("Mean",  "Arithmetic average of all samples in the region. "
                          "Indicates the typical load or baseline level."),
                ("Std Dev",
                 "Standard deviation of samples in the region. "
                 "Low Std Dev in a quasi-static region indicates a stable, noise-free hold. "
                 "High Std Dev in a dynamic region indicates a large or irregular transient."),
                ("Pk-Pk",
                 "Peak-to-peak amplitude (Max − Min). Represents the total swing of the signal "
                 "within the region, regardless of its sign or offset."),
            ]),
            ("Analysis controls", YELLOW, [
                ("Window",      "Smoothing window (seconds) applied to the absolute derivative "
                                "before thresholding. Larger values merge short bursts of "
                                "activity into a single dynamic region."),
                ("Threshold",   "Fraction (0–1) of the peak smoothed derivative used as the "
                                "dynamic/quasi-static boundary. Lower values classify more of "
                                "the signal as dynamic."),
                ("Min value",   "Regions whose peak amplitude is below this value are excluded "
                                "from the results entirely. Useful for ignoring noise-floor "
                                "artefacts or unloaded baseline segments."),
                ("Max regions", "Keeps only the N regions with the highest peak value after "
                                "all other filters are applied. Leave blank to show all regions."),
            ]),
        ]

        canvas = tk.Canvas(win, bg=BG, highlightthickness=0)
        scroll = tk.Scrollbar(win, orient=tk.VERTICAL, command=canvas.yview)
        canvas.configure(yscrollcommand=scroll.set)
        scroll.pack(side=tk.RIGHT, fill=tk.Y, padx=(0, 6), pady=6)
        canvas.pack(fill=tk.BOTH, expand=True, padx=(12, 0), pady=(0, 6))

        inner = tk.Frame(canvas, bg=BG)
        inner_id = canvas.create_window((0, 0), window=inner, anchor="nw")

        def _on_resize(e):
            canvas.itemconfig(inner_id, width=e.width)
        canvas.bind("<Configure>", _on_resize)
        inner.bind("<Configure>", lambda _: canvas.configure(
            scrollregion=canvas.bbox("all")))

        for section_title, colour, entries in sections:
            tk.Label(inner, text=section_title, font=("Helvetica", 10, "bold"),
                     fg=colour, bg=BG).pack(anchor="w", padx=10, pady=(10, 2))
            sep = tk.Frame(inner, bg=colour, height=1)
            sep.pack(fill=tk.X, padx=10, pady=(0, 6))

            for metric, description in entries:
                row = tk.Frame(inner, bg=SURFACE)
                row.pack(fill=tk.X, padx=10, pady=2)
                tk.Label(row, text=metric, font=("Courier New", 9, "bold"),
                         fg=colour, bg=SURFACE, width=14, anchor="nw",
                         justify=tk.LEFT).pack(side=tk.LEFT, padx=(8, 6), pady=6)
                tk.Label(row, text=description, font=("Helvetica", 9),
                         fg=TEXT, bg=SURFACE, wraplength=430,
                         justify=tk.LEFT, anchor="nw").pack(
                    side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 8), pady=6)

        tk.Button(win, text="Close", command=win.destroy,
                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=16, pady=4,
                  activebackground=MUTED, cursor="hand2",
                  font=("Helvetica", 9)).pack(pady=(4, 14))

    def _build_statusbar(self) -> None:
        bar = tk.Frame(self.root, bg=BG)
        bar.pack(side=tk.BOTTOM, fill=tk.X, padx=20, pady=(4, 10))
        self._status_var = tk.StringVar(value="Load an .xlsx recording to begin.")
        tk.Label(bar, textvariable=self._status_var, font=("Helvetica", 8),
                 fg=MUTED, bg=BG).pack(side=tk.LEFT)
        tk.Button(bar, text="Save filtered…", command=self._save,
                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=12, pady=3,
                  activebackground=MUTED, cursor="hand2").pack(side=tk.RIGHT)

    # ── File I/O ──────────────────────────────────────────────────────────────

    def _browse(self) -> None:
        path = filedialog.askopenfilename(
            title="Open recording",
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
        )
        if path:
            self._load_file(path)

    def _load_file(self, path: str) -> None:
        if not HAS_XLSX:
            messagebox.showerror("openpyxl missing", "pip install openpyxl")
            return
        try:
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception as exc:
            messagebox.showerror("Load failed", str(exc))
            return

        if len(rows) < 3:
            messagebox.showerror("Too short", "File needs at least 2 data rows.")
            return

        headers = [str(h) if h is not None else f"Col{i}" for i, h in enumerate(rows[0])]
        try:
            data = np.array([[float(c) if c is not None else np.nan for c in r]
                             for r in rows[1:]], dtype=float)
        except Exception as exc:
            messagebox.showerror("Parse error", str(exc))
            return

        if data.ndim < 2 or data.shape[1] < 2:
            messagebox.showerror("Format error", "Expected at least Time and one data column.")
            return

        self._time = data[:, 0]
        diffs = np.diff(self._time)
        diffs = diffs[diffs > 0]
        self._fs = 1.0 / np.median(diffs) if len(diffs) else 100.0
        nyq = self._fs / 2.0

        # Update slider limits to match this file's Nyquist frequency
        top = round(nyq * 0.99, 2)
        self._fc1_scale.configure(to=top)
        self._fc2_scale.configure(to=top)
        self._fc1_var.set(min(self._fc1_var.get(), nyq * 0.4))
        self._fc2_var.set(min(self._fc2_var.get(), nyq * 0.7))
        self._fc1_val_lbl.configure(text=f"{self._fc1_var.get():6.2f} Hz")
        self._fc2_val_lbl.configure(text=f"{self._fc2_var.get():6.2f} Hz")
        self._nyq_lbl.configure(text=f"Nyquist: {nyq:.2f} Hz")

        self._columns = {headers[i]: data[:, i] for i in range(1, len(headers))}
        cols = list(self._columns.keys())
        self._col_combo["values"] = cols
        if cols:
            self._col_var.set(cols[0])
            self._offset_unit_lbl.configure(text=cols[0])

        self._file_path = path
        self._path_var.set(pathlib.Path(path).name)
        n = len(self._time)
        dur = self._time[-1] - self._time[0]
        self._status_var.set(
            f"Loaded  {n:,} samples  |  fs ≈ {self._fs:.1f} Hz  |  "
            f"duration {dur:.2f} s  |  columns: {', '.join(cols)}")

        self._filtered = None
        self._apply_filter()

    # ── Filter ────────────────────────────────────────────────────────────────

    def _on_col_change(self, _=None) -> None:
        self._offset_unit_lbl.configure(text=self._col_var.get())
        self._schedule_apply()

    def _on_type_change(self, _=None) -> None:
        is_band = self._type_var.get() in ("Band-pass", "Band-stop")
        self._fc1_lbl_widget.configure(text="Low cutoff:" if is_band else "Cutoff:")
        if is_band:
            self._fc2_frame.pack(fill=tk.X, pady=(2, 0))
        else:
            self._fc2_frame.pack_forget()
        self._schedule_apply()

    def _schedule_apply(self, *_) -> None:
        if not self._auto_var.get():
            return
        if self._apply_job:
            self.root.after_cancel(self._apply_job)
        self._apply_job = self.root.after(250, self._apply_filter)

    def _apply_filter(self) -> None:
        if not HAS_SCIPY:
            messagebox.showerror("scipy missing",
                                 f"Install scipy:\n    pip install scipy"
                                 f"\n\nActual error:\n{SCIPY_ERR}")
            return
        if self._time is None or not self._columns:
            return

        col = self._col_var.get()
        if col not in self._columns:
            return
        raw = self._columns[col]
        nyq = self._fs / 2.0

        btype = BTYPE_MAP[self._type_var.get()]
        order = self._order_var.get()
        fc1   = float(self._fc1_var.get())

        if btype in ("bandpass", "bandstop"):
            fc2 = float(self._fc2_var.get())
            if fc2 <= fc1:
                self._status_var.set("Error: high cutoff must be greater than low cutoff.")
                return
            Wn = [fc1 / nyq, fc2 / nyq]
            fc_str = f"{fc1:.2f} – {fc2:.2f} Hz"
        else:
            Wn = fc1 / nyq
            fc_str = f"{fc1:.2f} Hz"

        # filtfilt needs at least 3*(2*order)+1 samples
        min_len = 3 * (2 * order) + 1
        if len(raw) < min_len:
            self._status_var.set(
                f"Error: need ≥ {min_len} samples for order-{order} filter "
                f"(have {len(raw)}).")
            return

        # Clamp Wn safely inside (0, 1) exclusive
        if isinstance(Wn, list):
            Wn = [max(1e-4, min(0.9999, w)) for w in Wn]
        else:
            Wn = max(1e-4, min(0.9999, Wn))

        # Remove DC offset before filtering so it doesn't cause filter artifacts
        input_sig = raw - np.mean(raw) if self._remove_dc_var.get() else raw

        try:
            b, a = butter(order, Wn, btype=btype)
            self._filtered = filtfilt(b, a, input_sig)
        except Exception as exc:
            self._status_var.set(f"Filter error: {exc}")
            return

        if self._invert_var.get():
            self._filtered = -self._filtered

        try:
            offset = float(self._offset_var.get())
        except ValueError:
            offset = 0.0
        if offset != 0.0:
            self._filtered = self._filtered + offset

        self._clear_analysis_spans()
        self._last_analysis = []

        # ── Signal plot ──────────────────────────────────────────────────────
        # Grey line shows input_sig (after DC removal, before filtering)
        # so the user can see exactly what the filter is acting on
        plot_input = -input_sig if self._invert_var.get() else input_sig
        self._plot_input = plot_input
        self._orig_line.set_data(self._time, plot_input)
        self._filt_line.set_data(self._time, self._filtered)
        self._sig_ax.set_ylabel(col, color=TEXT, fontsize=9)
        self._sig_ax.relim()
        self._sig_ax.autoscale_view()
        self._sig_canvas.draw_idle()

        # ── Frequency response ───────────────────────────────────────────────
        w, h = freqz(b, a, worN=4096, fs=self._fs)
        mag_db = 20.0 * np.log10(np.abs(h) + 1e-12)
        self._bode_line.set_data(w, mag_db)
        self._bode_ax.set_xlim(0.0, nyq)
        floor = max(-80.0, float(mag_db.min()) - 5.0)
        self._bode_ax.set_ylim(floor, 5.0)
        self._bode_canvas.draw_idle()

        extras = []
        if self._remove_dc_var.get():
            extras.append(f"DC removed (mean = {np.mean(raw):.4f})")
        if self._invert_var.get():
            extras.append("inverted")
        if offset != 0.0:
            extras.append(f"offset {offset:+.4g} {col}")
        extra_str = ("  |  " + ",  ".join(extras)) if extras else ""
        self._status_var.set(
            f"{self._type_var.get()}  |  order {order}  |  fc = {fc_str}  "
            f"|  '{col}'  |  {len(raw):,} samples{extra_str}")

    # ── Save ──────────────────────────────────────────────────────────────────

    def _save(self) -> None:
        if self._filtered is None:
            messagebox.showinfo("Nothing to save", "Apply a filter first.")
            return
        if not HAS_XLSX:
            messagebox.showerror("openpyxl missing", "pip install openpyxl")
            return

        default = pathlib.Path(self._file_path).stem + "_filtered.xlsx"
        path = filedialog.asksaveasfilename(
            title="Save filtered data",
            defaultextension=".xlsx",
            initialfile=default,
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
        )
        if not path:
            return

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Filtered"

        col_name   = self._col_var.get()
        all_cols   = ["Time (s)"] + list(self._columns.keys())
        out_hdrs   = all_cols + [f"{col_name} [filtered]"]

        hdr_font  = Font(bold=True, color="1E1E2E")
        hdr_fill  = PatternFill("solid", fgColor="89B4FA")
        hdr_align = Alignment(horizontal="center")
        for ci, h in enumerate(out_hdrs, start=1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = hdr_font
            c.fill = hdr_fill
            c.alignment = hdr_align

        col_names = list(self._columns.keys())
        for ri, t in enumerate(self._time):
            r = ri + 2
            ws.cell(row=r, column=1, value=round(float(t), 5))
            for ci, cname in enumerate(col_names, start=2):
                ws.cell(row=r, column=ci,
                        value=round(float(self._columns[cname][ri]), 6))
            ws.cell(row=r, column=len(all_cols) + 1,
                    value=round(float(self._filtered[ri]), 6))

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = (
                max(len(str(c.value or "")) for c in col) + 4)

        try:
            wb.save(path)
            self._status_var.set(f"Saved  →  {pathlib.Path(path).name}")
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))


if __name__ == "__main__":
    root = tk.Tk()
    FilterApp(root)
    root.mainloop()
