"""
NI USB-6002 + INA126 + Load Cell + Servo Monitor

AO0  -> 2.5 V reference voltage (held for process lifetime)
AI0  -> INA126 output measurement
USB  -> Arduino UNO via serial — Arduino generates precise servo PWM on pin 9
"""

import datetime
import json
import math
import pathlib
import threading
import time
from collections import deque
from dataclasses import asdict, dataclass
from typing import Optional

import numpy as np
import tkinter as tk
from tkinter import ttk, messagebox, filedialog

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_XLSX = True
except ImportError:
    HAS_XLSX = False

import matplotlib
matplotlib.use("TkAgg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg

try:
    import nidaqmx
    from nidaqmx.constants import TerminalConfiguration, AcquisitionType
    HAS_DAQ = True
except ImportError:
    HAS_DAQ = False

try:
    import serial
    import serial.tools.list_ports
    HAS_SERIAL = True
except ImportError:
    HAS_SERIAL = False

# ── Hardware ────────────────────────────────────────────────────────────────
DEVICE_NAME       = "Dev1"
AI_CHANNEL        = f"{DEVICE_NAME}/ai0"
AO_CHANNEL        = f"{DEVICE_NAME}/ao0"
REFERENCE_VOLTAGE = 2.5
AI_MIN_V          = -10.0
AI_MAX_V          =  10.0
SAMPLE_RATE       = 100
CHUNK_SIZE        = 10
SERVO_MIN_ANGLE   = 0
SERVO_MAX_ANGLE   = 180

# ── Arduino serial ───────────────────────────────────────────────────────────
SERIAL_PORT = "COM3"    # change to match your Arduino's port (check Device Manager)
SERIAL_BAUD = 9600

# ── Display / calibration ───────────────────────────────────────────────────
HISTORY_SECONDS  = 15
MAX_POINTS       = SAMPLE_RATE * HISTORY_SECONDS
CAL_AVG_SAMPLES  = 100
CALIBRATION_FILE = pathlib.Path(__file__).with_name("calibration.json")
CAL_UNITS        = ["kg", "g", "lb", "oz", "N"]
GRAPH_UNITS      = ["Voltage"] + CAL_UNITS

# Conversion factors: how many of each unit equal 1 kg
KG_TO_UNIT = {"kg": 1.0, "g": 1000.0, "lb": 2.20462, "oz": 35.274, "N": 9.80665}

# ── Colours ─────────────────────────────────────────────────────────────────
BG      = "#1e1e2e"
SURFACE = "#313244"
MUTED   = "#6c7086"
TEXT    = "#cdd6f4"
BLUE    = "#89b4fa"
GREEN   = "#a6e3a1"
YELLOW  = "#f9e2af"
MAUVE   = "#cba6f7"
BORDER  = "#45475a"



# ── Calibration model ───────────────────────────────────────────────────────

@dataclass
class Calibration:
    zero_v:       Optional[float] = None
    span_v:       Optional[float] = None
    known_weight: Optional[float] = None
    unit:         str             = "kg"

    def is_valid(self) -> bool:
        return (
            self.zero_v is not None
            and self.span_v is not None
            and self.known_weight is not None
            and abs(self.span_v - self.zero_v) > 1e-6
        )

    def to_weight(self, voltage: float) -> Optional[float]:
        if not self.is_valid():
            return None
        return (voltage - self.zero_v) / (self.span_v - self.zero_v) * self.known_weight

    def to_weight_in_unit(self, voltage: float, display_unit: str) -> Optional[float]:
        w = self.to_weight(voltage)
        if w is None:
            return None
        if display_unit == self.unit:
            return w
        w_kg = w / KG_TO_UNIT[self.unit]
        return w_kg * KG_TO_UNIT[display_unit]

    def sensitivity_str(self) -> str:
        if not self.is_valid():
            return "—"
        sens = self.known_weight / (self.span_v - self.zero_v)
        return f"{sens:.4f} {self.unit}/V"

    def save(self) -> None:
        CALIBRATION_FILE.write_text(json.dumps(asdict(self), indent=2))

    def load(self) -> bool:
        if not CALIBRATION_FILE.exists():
            return False
        try:
            d = json.loads(CALIBRATION_FILE.read_text())
            self.zero_v       = d.get("zero_v")
            self.span_v       = d.get("span_v")
            self.known_weight = d.get("known_weight")
            self.unit         = d.get("unit", "kg")
            return True
        except Exception:
            return False

    def clear(self) -> None:
        self.zero_v = self.span_v = self.known_weight = None


# ── Calibration dialog ──────────────────────────────────────────────────────

class CalibrationDialog(tk.Toplevel):
    def __init__(self, parent: tk.Tk, app: "App") -> None:
        super().__init__(parent)
        self._app = app
        self.title("Load Cell Calibration")
        self.configure(bg=BG)
        self.geometry("500x360")
        self.resizable(False, False)
        self.transient(parent)
        self.grab_set()
        self._build()
        self._refresh()

    def _build(self) -> None:
        tk.Label(self, text="Load Cell Calibration",
                 font=("Helvetica", 14, "bold"), fg=TEXT, bg=BG).pack(pady=(16, 2))
        tk.Label(self, text="Two-point linear calibration — follow steps in order.",
                 font=("Helvetica", 9), fg=MUTED, bg=BG).pack(pady=(0, 10))

        # Step 1
        s1 = tk.Frame(self, bg=SURFACE)
        s1.pack(fill=tk.X, padx=20, pady=(0, 6))
        tk.Label(s1, text="Step 1 — Tare (zero)", font=("Helvetica", 9, "bold"),
                 fg=YELLOW, bg=SURFACE).pack(anchor="w", padx=10, pady=(8, 2))
        tk.Label(s1, text="Remove all weight from the load cell, then click Tare.",
                 font=("Helvetica", 8), fg=MUTED, bg=SURFACE).pack(anchor="w", padx=10)
        r1 = tk.Frame(s1, bg=SURFACE)
        r1.pack(fill=tk.X, padx=10, pady=(6, 10))
        tk.Button(r1, text="Tare / Zero", command=self._tare,
                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=14, pady=5,
                  activebackground=MUTED, cursor="hand2").pack(side=tk.LEFT)
        self._zero_lbl = tk.Label(r1, text="not set", font=("Courier New", 10),
                                  fg=MUTED, bg=SURFACE)
        self._zero_lbl.pack(side=tk.LEFT, padx=14)

        # Step 2
        s2 = tk.Frame(self, bg=SURFACE)
        s2.pack(fill=tk.X, padx=20, pady=(0, 6))
        tk.Label(s2, text="Step 2 — Set span (known weight)",
                 font=("Helvetica", 9, "bold"), fg=YELLOW, bg=SURFACE).pack(anchor="w", padx=10, pady=(8, 2))
        tk.Label(s2, text="Apply a known weight, enter the value below, then click Set Span.",
                 font=("Helvetica", 8), fg=MUTED, bg=SURFACE).pack(anchor="w", padx=10)
        r2 = tk.Frame(s2, bg=SURFACE)
        r2.pack(fill=tk.X, padx=10, pady=(6, 10))
        tk.Label(r2, text="Weight:", font=("Helvetica", 9), fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        cal = self._app._cal
        self._weight_var = tk.StringVar(
            value=str(cal.known_weight) if cal.known_weight is not None else "")
        tk.Entry(r2, textvariable=self._weight_var, width=10,
                 bg=BORDER, fg=TEXT, insertbackground=TEXT,
                 relief=tk.FLAT, font=("Courier New", 11)).pack(side=tk.LEFT, padx=6)
        self._unit_var = tk.StringVar(value=cal.unit)
        ttk.Combobox(r2, textvariable=self._unit_var,
                     values=CAL_UNITS, width=5, state="readonly").pack(side=tk.LEFT, padx=(0, 14))
        tk.Button(r2, text="Set Span", command=self._set_span,
                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=14, pady=5,
                  activebackground=MUTED, cursor="hand2").pack(side=tk.LEFT)
        self._span_lbl = tk.Label(r2, text="not set", font=("Courier New", 10),
                                  fg=MUTED, bg=SURFACE)
        self._span_lbl.pack(side=tk.LEFT, padx=14)

        info = tk.Frame(self, bg=BG)
        info.pack(fill=tk.X, padx=20, pady=(4, 0))
        self._sens_lbl = tk.Label(info, text="", font=("Helvetica", 8), fg=MUTED, bg=BG)
        self._sens_lbl.pack(side=tk.LEFT)
        self._status_lbl = tk.Label(info, text="", font=("Helvetica", 8, "bold"), fg=GREEN, bg=BG)
        self._status_lbl.pack(side=tk.RIGHT)

        btns = tk.Frame(self, bg=BG)
        btns.pack(fill=tk.X, padx=20, pady=(10, 16))

        def _btn(parent, text, cmd, color=BORDER):
            return tk.Button(parent, text=text, command=cmd,
                             bg=color, fg=TEXT, relief=tk.FLAT, padx=12, pady=4,
                             activebackground=MUTED, cursor="hand2")

        _btn(btns, "Load saved", self._load_cal).pack(side=tk.LEFT, padx=(0, 4))
        _btn(btns, "Clear", self._clear_cal, "#592929").pack(side=tk.LEFT)
        _btn(btns, "Close", self.destroy).pack(side=tk.RIGHT)

    def _snapshot(self) -> Optional[float]:
        with self._app._lock:
            recent = list(self._app._voltage_buf)[-CAL_AVG_SAMPLES:]
        if len(recent) < 5:
            messagebox.showwarning("No data",
                                   "Wait for readings to appear before calibrating.",
                                   parent=self)
            return None
        return float(np.mean(recent))

    def _tare(self) -> None:
        v = self._snapshot()
        if v is None:
            return
        self._app._cal.zero_v = v
        self._app._cal.save()
        self._flash(f"Zero captured: {v:.5f} V  (saved)")
        self._refresh()

    def _set_span(self) -> None:
        if self._app._cal.zero_v is None:
            messagebox.showwarning("Tare first",
                                   "Complete Step 1 (Tare) before setting the span.",
                                   parent=self)
            return
        try:
            known = float(self._weight_var.get().strip())
            if known <= 0:
                raise ValueError
        except ValueError:
            messagebox.showerror("Invalid weight",
                                 "Enter a positive number for the known weight.",
                                 parent=self)
            return
        v = self._snapshot()
        if v is None:
            return
        if abs(v - self._app._cal.zero_v) < 1e-5:
            messagebox.showwarning("No deflection",
                                   "Reading is too close to the zero point.\n"
                                   "Make sure the weight is applied before clicking Set Span.",
                                   parent=self)
            return
        self._app._cal.span_v       = v
        self._app._cal.known_weight = known
        self._app._cal.unit         = self._unit_var.get()
        self._app._cal.save()
        self._flash(f"Span set: {v:.5f} V = {known} {self._app._cal.unit}  (saved)")
        self._refresh()
        self._app._refresh_weight_display()

    def _load_cal(self) -> None:
        if self._app._cal.load():
            self._unit_var.set(self._app._cal.unit)
            if self._app._cal.known_weight is not None:
                self._weight_var.set(str(self._app._cal.known_weight))
            self._refresh()
            self._app._refresh_weight_display()
            self._flash(f"Loaded  ←  {CALIBRATION_FILE.name}")
        else:
            messagebox.showinfo("Not found",
                                f"{CALIBRATION_FILE.name} does not exist yet.",
                                parent=self)

    def _clear_cal(self) -> None:
        if not messagebox.askyesno("Clear calibration",
                                   "This will clear the current calibration.\nContinue?",
                                   parent=self):
            return
        self._app._cal.clear()
        self._refresh()
        self._app._refresh_weight_display()
        self._flash("Calibration cleared", YELLOW)

    def _refresh(self) -> None:
        cal = self._app._cal
        if cal.zero_v is not None:
            self._zero_lbl.configure(text=f"{cal.zero_v:.5f} V  ✓", fg=GREEN)
        else:
            self._zero_lbl.configure(text="not set", fg=MUTED)
        if cal.span_v is not None and cal.known_weight is not None:
            self._span_lbl.configure(
                text=f"{cal.span_v:.5f} V = {cal.known_weight} {cal.unit}  ✓", fg=GREEN)
        else:
            self._span_lbl.configure(text="not set", fg=MUTED)
        self._sens_lbl.configure(text=f"Sensitivity:  {cal.sensitivity_str()}")

    def _flash(self, msg: str, color: str = GREEN) -> None:
        self._status_lbl.configure(text=msg, fg=color)
        self.after(3000, lambda: self._status_lbl.configure(text=""))


# ── Main application ────────────────────────────────────────────────────────

class App:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("INA126 / NI USB-6002 Monitor")
        self.root.geometry("980x980")
        self.root.configure(bg=BG)
        self.root.resizable(True, True)

        # Measurement buffers
        self._voltage_buf: deque[float] = deque(maxlen=MAX_POINTS)
        self._time_buf:    deque[float] = deque(maxlen=MAX_POINTS)
        self._t0      = time.time()
        self._lock    = threading.Lock()
        self._running = False
        self._ao_task = None
        self._after_id: Optional[str] = None

        # Calibration
        self._cal = Calibration()
        self._cal.load()
        self._graph_unit_var = tk.StringVar(value="Voltage")

        # Recording
        self._recording    = False
        self._record_buf: list = []
        self._record_start = 0.0
        self._record_path  = ""

        # Servo / Arduino serial
        self._servo_angle      = 90.0
        self._last_sent_angle  = -1
        self._serial_port: Optional[serial.Serial] = None if HAS_SERIAL else None
        self._servo_angle_buf: deque[float]  = deque(maxlen=MAX_POINTS)
        self._servo_time_buf:  deque[float]  = deque(maxlen=MAX_POINTS)

        # Waveform generator
        self._waveform_active  = False

        self._build_ui()
        self._refresh_weight_display()
        self._start()

    # ── UI ───────────────────────────────────────────────────────────────────

    def _build_ui(self) -> None:
        self._build_header()
        self._build_readout()
        self._build_stats()

        # Container that holds both graphs + servo panel, sharing vertical space
        graphs_outer = tk.Frame(self.root, bg=BG)
        graphs_outer.pack(fill=tk.BOTH, expand=True, padx=20, pady=(8, 0))

        self._build_signal_graph(graphs_outer)
        self._build_servo_panel(graphs_outer)
        self._build_angle_graph(graphs_outer)

        self._build_statusbar()
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

    def _build_header(self) -> None:
        hdr = tk.Frame(self.root, bg=BG)
        hdr.pack(fill=tk.X, padx=20, pady=(12, 0))
        tk.Label(hdr, text="INA126 / Servo Monitor", font=("Helvetica", 16, "bold"),
                 fg=TEXT, bg=BG).pack(side=tk.LEFT)
        tk.Button(hdr, text="Calibrate…", command=self._open_calibration,
                  bg=YELLOW, fg="#1e1e2e", font=("Helvetica", 10, "bold"),
                  relief=tk.FLAT, padx=12, pady=4, cursor="hand2").pack(side=tk.RIGHT)
        tk.Label(hdr, text=f"REF  {REFERENCE_VOLTAGE} V  on  AO0",
                 font=("Helvetica", 11), fg=GREEN, bg=BG).pack(side=tk.RIGHT, padx=16)

    def _build_readout(self) -> None:
        ro = tk.Frame(self.root, bg=SURFACE)
        ro.pack(fill=tk.X, padx=20, pady=10)
        ro.columnconfigure(0, weight=1)
        ro.columnconfigure(1, weight=1)

        vf = tk.Frame(ro, bg=SURFACE)
        vf.grid(row=0, column=0, sticky="nsew", padx=16, pady=10)
        tk.Label(vf, text="VOLTAGE  (V)", font=("Helvetica", 9), fg=MUTED, bg=SURFACE).pack()
        self._v_var = tk.StringVar(value="---")
        tk.Label(vf, textvariable=self._v_var, font=("Courier New", 42, "bold"),
                 fg=BLUE, bg=SURFACE).pack()

        tk.Frame(ro, bg=BORDER, width=1).grid(row=0, column=0, sticky="nse", pady=8)

        wf = tk.Frame(ro, bg=SURFACE)
        wf.grid(row=0, column=1, sticky="nsew", padx=16, pady=10)
        self._w_unit_var = tk.StringVar(value=f"WEIGHT  ({self._cal.unit})")
        tk.Label(wf, textvariable=self._w_unit_var, font=("Helvetica", 9), fg=MUTED, bg=SURFACE).pack()
        self._w_var = tk.StringVar(value="---")
        self._w_lbl = tk.Label(wf, textvariable=self._w_var,
                               font=("Courier New", 42, "bold"), fg=MUTED, bg=SURFACE)
        self._w_lbl.pack()

    def _build_stats(self) -> None:
        stats = tk.Frame(self.root, bg=BG)
        stats.pack(fill=tk.X, padx=20)
        self._stat_vars: dict[str, tk.StringVar] = {}
        for col, key in enumerate(("MIN", "MAX", "AVG")):
            f = tk.Frame(stats, bg=SURFACE)
            f.grid(row=0, column=col, padx=(0 if col == 0 else 6, 0), sticky="ew")
            stats.columnconfigure(col, weight=1)
            tk.Label(f, text=key, font=("Helvetica", 8), fg=MUTED, bg=SURFACE).pack(pady=(4, 0))
            var = tk.StringVar(value="---")
            self._stat_vars[key] = var
            tk.Label(f, textvariable=var, font=("Helvetica", 13, "bold"),
                     fg=TEXT, bg=SURFACE).pack(pady=(0, 4))

    def _build_signal_graph(self, parent: tk.Frame) -> None:
        gf = tk.Frame(parent, bg=BG)
        gf.pack(fill=tk.BOTH, expand=True)

        # Y-axis selector
        sel = tk.Frame(gf, bg=BG)
        sel.pack(fill=tk.X, pady=(4, 2))
        tk.Label(sel, text="Y axis:", font=("Helvetica", 8),
                 fg=MUTED, bg=BG).pack(side=tk.LEFT, padx=(0, 6))
        for unit in GRAPH_UNITS:
            tk.Radiobutton(
                sel, text=unit, variable=self._graph_unit_var, value=unit,
                command=self._on_graph_unit_change,
                bg=BG, fg=TEXT, selectcolor=SURFACE,
                activebackground=BG, activeforeground=TEXT,
                font=("Helvetica", 8), cursor="hand2",
            ).pack(side=tk.LEFT, padx=3)

        self._fig, self._ax = plt.subplots(figsize=(9, 2.6))
        self._fig.patch.set_facecolor(BG)
        self._ax.set_facecolor(SURFACE)
        self._ax.tick_params(colors=TEXT, labelsize=8)
        for spine in self._ax.spines.values():
            spine.set_color(BORDER)
        self._ax.set_ylabel("Voltage (V)", color=TEXT, fontsize=9)
        self._ax.set_xlabel("Time (s)", color=TEXT, fontsize=9)
        self._ax.grid(color=BORDER, linewidth=0.5, alpha=0.6)
        self._line, = self._ax.plot([], [], color=BLUE, linewidth=1.5, antialiased=True)
        self._ref_line = self._ax.axhline(
            y=REFERENCE_VOLTAGE, color=GREEN, linestyle="--",
            linewidth=1.0, alpha=0.8, label=f"{REFERENCE_VOLTAGE} V ref")
        self._ax.legend(facecolor=SURFACE, edgecolor=BORDER, labelcolor=TEXT, fontsize=8)
        self._fig.tight_layout(pad=1.0)

        self._canvas = FigureCanvasTkAgg(self._fig, master=gf)
        self._canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def _build_servo_panel(self, parent: tk.Frame) -> None:
        panel = tk.LabelFrame(parent, text="  SERVO CONTROL  ",
                              font=("Helvetica", 9, "bold"),
                              fg=MAUVE, bg=SURFACE, bd=1, relief=tk.FLAT,
                              labelanchor="nw")
        panel.pack(fill=tk.X, pady=(6, 0))

        # ── Row 0: slider + angle display ──
        r0 = tk.Frame(panel, bg=SURFACE)
        r0.pack(fill=tk.X, padx=10, pady=(6, 4))

        tk.Label(r0, text="Angle:", font=("Helvetica", 9),
                 fg=TEXT, bg=SURFACE).pack(side=tk.LEFT, padx=(0, 6))

        self._servo_scale_var = tk.DoubleVar(value=self._servo_angle)
        self._servo_scale = tk.Scale(
            r0, from_=SERVO_MIN_ANGLE, to=SERVO_MAX_ANGLE,
            orient=tk.HORIZONTAL, variable=self._servo_scale_var,
            command=self._on_slider_change,
            bg=SURFACE, fg=TEXT, troughcolor=BORDER,
            highlightthickness=0, activebackground=MAUVE,
            length=300, showvalue=False,
        )
        self._servo_scale.pack(side=tk.LEFT)

        self._servo_disp_var = tk.StringVar(value="90.0°")
        tk.Label(r0, textvariable=self._servo_disp_var,
                 font=("Courier New", 18, "bold"), fg=MAUVE, bg=SURFACE,
                 width=7).pack(side=tk.LEFT, padx=10)

        # COM port selector
        port_frame = tk.Frame(r0, bg=SURFACE)
        port_frame.pack(side=tk.RIGHT)
        tk.Label(port_frame, text="Arduino port:", font=("Helvetica", 8),
                 fg=MUTED, bg=SURFACE).pack(side=tk.LEFT, padx=(0, 4))
        self._port_var = tk.StringVar(value=SERIAL_PORT)
        self._port_combo = ttk.Combobox(port_frame, textvariable=self._port_var,
                                        width=8, font=("Helvetica", 8))
        self._port_combo.pack(side=tk.LEFT)
        self._port_combo.bind("<Button-1>", self._refresh_ports)
        tk.Button(port_frame, text="Connect", command=self._connect_serial,
                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=8, pady=1,
                  activebackground=MAUVE, cursor="hand2",
                  font=("Helvetica", 8)).pack(side=tk.LEFT, padx=(4, 0))
        self._serial_status_var = tk.StringVar(value="disconnected")
        self._serial_status_lbl = tk.Label(port_frame, textvariable=self._serial_status_var,
                                           font=("Helvetica", 8), fg="#f38ba8", bg=SURFACE)
        self._serial_status_lbl.pack(side=tk.LEFT, padx=(6, 0))

        # ── Row 1: preset buttons ──
        r1 = tk.Frame(panel, bg=SURFACE)
        r1.pack(fill=tk.X, padx=10, pady=(0, 6))
        tk.Label(r1, text="Presets:", font=("Helvetica", 8),
                 fg=MUTED, bg=SURFACE).pack(side=tk.LEFT, padx=(0, 6))
        self._preset_btns: list[tk.Button] = []
        for deg in (0, 45, 90, 135, 180):
            btn = tk.Button(
                r1, text=f"{deg}°", width=4,
                command=lambda d=deg: self._set_servo_angle(d),
                bg=BORDER, fg=TEXT, relief=tk.FLAT, pady=2,
                activebackground=MAUVE, activeforeground=BG, cursor="hand2",
            )
            btn.pack(side=tk.LEFT, padx=3)
            self._preset_btns.append(btn)

        # ── Row 2: waveform generator ──
        wf = tk.LabelFrame(panel, text="  WAVEFORM  ",
                           font=("Helvetica", 8, "bold"),
                           fg=BLUE, bg=SURFACE, bd=1, relief=tk.FLAT,
                           labelanchor="nw")
        wf.pack(fill=tk.X, padx=10, pady=(2, 10))

        # Sub-row A: type + frequency + start/stop
        wa = tk.Frame(wf, bg=SURFACE)
        wa.pack(fill=tk.X, padx=8, pady=(6, 2))

        tk.Label(wa, text="Shape:", font=("Helvetica", 8), fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._wave_type_var = tk.StringVar(value="Sine")
        ttk.Combobox(wa, textvariable=self._wave_type_var, state="readonly", width=14,
                     values=["Sine", "Square", "Triangle", "Sawtooth", "Rev. Sawtooth"]
                     ).pack(side=tk.LEFT, padx=(4, 16))

        tk.Label(wa, text="Freq:", font=("Helvetica", 8), fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._wave_freq_var = tk.DoubleVar(value=0.5)
        self._wave_freq_lbl = tk.Label(wa, text="0.50 Hz", font=("Courier New", 9),
                                       fg=BLUE, bg=SURFACE, width=8)
        tk.Scale(wa, from_=0.05, to=5.0, resolution=0.05, orient=tk.HORIZONTAL,
                 variable=self._wave_freq_var, length=160, showvalue=False,
                 bg=SURFACE, fg=TEXT, troughcolor=BORDER, highlightthickness=0,
                 activebackground=BLUE,
                 command=lambda v: self._wave_freq_lbl.configure(text=f"{float(v):.2f} Hz")
                 ).pack(side=tk.LEFT, padx=(4, 2))
        self._wave_freq_lbl.pack(side=tk.LEFT, padx=(0, 16))

        self._wave_btn = tk.Button(wa, text="▶  Start", width=10,
                                   command=self._toggle_waveform,
                                   bg=BORDER, fg=TEXT, relief=tk.FLAT, pady=3,
                                   activebackground=BLUE, cursor="hand2",
                                   font=("Helvetica", 9, "bold"))
        self._wave_btn.pack(side=tk.RIGHT)

        # Sub-row B: center + amplitude
        wb = tk.Frame(wf, bg=SURFACE)
        wb.pack(fill=tk.X, padx=8, pady=(2, 8))

        tk.Label(wb, text="Center:", font=("Helvetica", 8), fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._wave_center_var = tk.DoubleVar(value=90.0)
        self._wave_center_lbl = tk.Label(wb, text=" 90°", font=("Courier New", 9),
                                         fg=BLUE, bg=SURFACE, width=5)
        tk.Scale(wb, from_=0, to=180, resolution=1, orient=tk.HORIZONTAL,
                 variable=self._wave_center_var, length=160, showvalue=False,
                 bg=SURFACE, fg=TEXT, troughcolor=BORDER, highlightthickness=0,
                 activebackground=BLUE,
                 command=lambda v: self._wave_center_lbl.configure(text=f"{int(float(v)):3d}°")
                 ).pack(side=tk.LEFT, padx=(4, 2))
        self._wave_center_lbl.pack(side=tk.LEFT, padx=(0, 20))

        tk.Label(wb, text="Amplitude:", font=("Helvetica", 8), fg=TEXT, bg=SURFACE).pack(side=tk.LEFT)
        self._wave_amp_var = tk.DoubleVar(value=45.0)
        self._wave_amp_lbl = tk.Label(wb, text=" 45°", font=("Courier New", 9),
                                      fg=BLUE, bg=SURFACE, width=5)
        tk.Scale(wb, from_=0, to=90, resolution=1, orient=tk.HORIZONTAL,
                 variable=self._wave_amp_var, length=160, showvalue=False,
                 bg=SURFACE, fg=TEXT, troughcolor=BORDER, highlightthickness=0,
                 activebackground=BLUE,
                 command=lambda v: self._wave_amp_lbl.configure(text=f"{int(float(v)):3d}°")
                 ).pack(side=tk.LEFT, padx=(4, 2))
        self._wave_amp_lbl.pack(side=tk.LEFT)

    def _build_angle_graph(self, parent: tk.Frame) -> None:
        gf = tk.Frame(parent, bg=BG)
        gf.pack(fill=tk.BOTH, expand=True, pady=(6, 0))

        lbl = tk.Frame(gf, bg=BG)
        lbl.pack(fill=tk.X)
        tk.Label(lbl, text="Servo Angle", font=("Helvetica", 8),
                 fg=MUTED, bg=BG).pack(side=tk.LEFT)

        self._angle_fig, self._angle_ax = plt.subplots(figsize=(9, 2.2))
        self._angle_fig.patch.set_facecolor(BG)
        self._angle_ax.set_facecolor(SURFACE)
        self._angle_ax.tick_params(colors=TEXT, labelsize=8)
        for spine in self._angle_ax.spines.values():
            spine.set_color(BORDER)
        self._angle_ax.set_ylabel("Angle (°)", color=TEXT, fontsize=9)
        self._angle_ax.set_xlabel("Time (s)", color=TEXT, fontsize=9)
        self._angle_ax.set_ylim(-5, 185)
        self._angle_ax.set_yticks([0, 45, 90, 135, 180])
        self._angle_ax.grid(color=BORDER, linewidth=0.5, alpha=0.6)
        self._angle_line, = self._angle_ax.plot([], [], color=MAUVE,
                                                 linewidth=1.5, antialiased=True)
        self._angle_fig.tight_layout(pad=1.0)

        self._angle_canvas = FigureCanvasTkAgg(self._angle_fig, master=gf)
        self._angle_canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

    def _build_statusbar(self) -> None:
        bar = tk.Frame(self.root, bg=BG)
        bar.pack(fill=tk.X, padx=20, pady=(4, 10))
        self._status_var = tk.StringVar(value="Initialising …")
        tk.Label(bar, textvariable=self._status_var, font=("Helvetica", 8),
                 fg=MUTED, bg=BG).pack(side=tk.LEFT)
        tk.Button(bar, text="Clear", command=self._clear,
                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=12, pady=2,
                  activebackground=MUTED).pack(side=tk.RIGHT)
        self._rec_count_var = tk.StringVar(value="")
        tk.Label(bar, textvariable=self._rec_count_var, font=("Helvetica", 8),
                 fg="#f38ba8", bg=BG).pack(side=tk.RIGHT, padx=8)
        self._rec_btn = tk.Button(bar, text="● Record", command=self._toggle_recording,
                                  bg=BORDER, fg=TEXT, relief=tk.FLAT, padx=12, pady=2,
                                  activebackground=MUTED, cursor="hand2")
        self._rec_btn.pack(side=tk.RIGHT, padx=(0, 6))

    # ── Calibration ───────────────────────────────────────────────────────────

    def _open_calibration(self) -> None:
        CalibrationDialog(self.root, self)

    def _on_graph_unit_change(self) -> None:
        unit = self._graph_unit_var.get()
        if unit != "Voltage" and not self._cal.is_valid():
            messagebox.showwarning("Not calibrated",
                                   "Calibrate the load cell before selecting a weight unit.")
            self._graph_unit_var.set("Voltage")
            return
        self._ref_line.set_visible(unit == "Voltage")
        self._ax.set_ylabel("Voltage (V)" if unit == "Voltage" else unit,
                            color=TEXT, fontsize=9)

    def _refresh_weight_display(self) -> None:
        if self._cal.is_valid():
            self._w_unit_var.set(f"WEIGHT  ({self._cal.unit})")
            self._w_lbl.configure(fg=GREEN)
        else:
            self._w_unit_var.set("WEIGHT")
            self._w_var.set("not calibrated")
            self._w_lbl.configure(fg=MUTED)

    # ── Servo / Arduino serial ────────────────────────────────────────────────

    def _refresh_ports(self, *_) -> None:
        if not HAS_SERIAL:
            return
        ports = [p.device for p in serial.tools.list_ports.comports()]
        self._port_combo["values"] = ports

    def _connect_serial(self) -> None:
        if not HAS_SERIAL:
            messagebox.showerror("pyserial not installed",
                                 "Run:  pip install pyserial\nthen restart the app.")
            return
        self._stop_servo()
        port = self._port_var.get().strip()
        try:
            self._serial_port = serial.Serial(port, SERIAL_BAUD, timeout=1)
            time.sleep(2)           # wait for Arduino to finish its reset after DTR toggle
            self._serial_status_var.set("connected")
            self._serial_status_lbl.configure(fg=GREEN)
            # Send current angle immediately so servo moves to the slider position
            self._send_angle(int(round(self._servo_angle)))
        except Exception as exc:
            self._serial_port = None
            self._serial_status_var.set("failed")
            messagebox.showerror("Serial error", str(exc))

    def _send_angle(self, angle_int: int) -> None:
        if self._serial_port is not None and self._serial_port.is_open:
            try:
                self._serial_port.write(f"{angle_int}\n".encode())
            except Exception:
                self._serial_port = None
                self._serial_status_var.set("disconnected")
                self._serial_status_lbl.configure(fg="#f38ba8")

    def _set_servo_angle(self, angle: float) -> None:
        angle = max(float(SERVO_MIN_ANGLE), min(float(SERVO_MAX_ANGLE), float(angle)))
        self._servo_angle = angle
        self._servo_scale_var.set(angle)
        self._servo_disp_var.set(f"{angle:.1f}°")
        angle_int = int(round(angle))
        if angle_int != self._last_sent_angle:
            self._last_sent_angle = angle_int
            self._send_angle(angle_int)

    def _on_slider_change(self, val: str) -> None:
        if not self._waveform_active:
            self._set_servo_angle(float(val))

    def _stop_servo(self) -> None:
        if self._serial_port is not None:
            try:
                self._serial_port.close()
            except Exception:
                pass
            self._serial_port = None
            self._serial_status_var.set("disconnected")

    # ── Waveform generator ────────────────────────────────────────────────────

    def _toggle_waveform(self) -> None:
        if self._waveform_active:
            self._waveform_active = False
        else:
            self._waveform_active = True
            self._servo_scale.configure(state="disabled")
            for btn in self._preset_btns:
                btn.configure(state="disabled")
            self._wave_btn.configure(text="■  Stop", bg="#592929")
            threading.Thread(target=self._waveform_thread_fn, daemon=True).start()

    def _on_waveform_stopped(self) -> None:
        self._servo_scale.configure(state="normal")
        for btn in self._preset_btns:
            btn.configure(state="normal")
        self._wave_btn.configure(text="▶  Start", bg=BORDER)

    @staticmethod
    def _compute_wave(wave_type: str, phase: float) -> float:
        """Return a value in [-1, 1] for the given waveform shape at phase [0, 1)."""
        if wave_type == "Sine":
            return math.sin(2.0 * math.pi * phase)
        if wave_type == "Square":
            return 1.0 if phase < 0.5 else -1.0
        if wave_type == "Triangle":
            # -1 at 0, +1 at 0.5, -1 at 1
            return 1.0 - 2.0 * abs(2.0 * phase - 1.0)
        if wave_type == "Sawtooth":
            return 2.0 * phase - 1.0          # ramps -1 → +1
        if wave_type == "Rev. Sawtooth":
            return 1.0 - 2.0 * phase          # ramps +1 → -1
        return 0.0

    def _waveform_thread_fn(self) -> None:
        UPDATE_HZ = 25
        interval  = 1.0 / UPDATE_HZ
        t_start   = time.perf_counter()
        t_next    = t_start + interval

        while self._waveform_active:
            t     = time.perf_counter() - t_start
            freq  = self._wave_freq_var.get()
            amp   = self._wave_amp_var.get()
            ctr   = self._wave_center_var.get()

            phase = (t * freq) % 1.0
            raw   = self._compute_wave(self._wave_type_var.get(), phase)
            angle = max(0.0, min(180.0, ctr + amp * raw))

            self._servo_angle = angle
            self._send_angle(int(round(angle)))

            now   = time.perf_counter()
            sleep = t_next - now
            if sleep > 0.001:
                time.sleep(sleep)
            t_next += interval

        self.root.after(0, self._on_waveform_stopped)

    # ── Hardware ─────────────────────────────────────────────────────────────

    def _start(self) -> None:
        self._running = True
        threading.Thread(target=self._hardware_thread, daemon=True).start()
        self._tick()

    def _hardware_thread(self) -> None:
        if not HAS_DAQ:
            self._status_var.set("nidaqmx not installed — running in DEMO mode")
            self._demo_loop()
            return

        try:
            self._ao_task = nidaqmx.Task()
            self._ao_task.ao_channels.add_ao_voltage_chan(
                AO_CHANNEL, min_val=0.0, max_val=5.0)
            self._ao_task.write(REFERENCE_VOLTAGE)
            self._ao_task.start()
        except Exception as exc:
            self._status_var.set(f"AO warning: {exc}")

        try:
            with nidaqmx.Task() as ai:
                ai.ai_channels.add_ai_voltage_chan(
                    AI_CHANNEL,
                    terminal_config=TerminalConfiguration.RSE,
                    min_val=AI_MIN_V,
                    max_val=AI_MAX_V,
                )
                ai.timing.cfg_samp_clk_timing(
                    rate=SAMPLE_RATE,
                    sample_mode=AcquisitionType.CONTINUOUS,
                    samps_per_chan=CHUNK_SIZE,
                )
                ai.start()
                self._status_var.set(
                    f"{AI_CHANNEL}  @  {SAMPLE_RATE} Sa/s  |  "
                    f"AO0 = {REFERENCE_VOLTAGE} V  |  Arduino servo on {self._port_var.get()}")
                while self._running:
                    samples = ai.read(number_of_samples_per_channel=CHUNK_SIZE)
                    t_now = time.time() - self._t0
                    with self._lock:
                        for i, v in enumerate(samples):
                            t = t_now - (CHUNK_SIZE - 1 - i) / SAMPLE_RATE
                            self._voltage_buf.append(float(v))
                            self._time_buf.append(t)
                            if self._recording:
                                w = self._cal.to_weight(float(v)) if self._cal.is_valid() else None
                                self._record_buf.append(
                                    (t - self._record_start, float(v), w, self._servo_angle))
        except Exception as exc:
            self._status_var.set(f"AI error: {exc} — falling back to DEMO mode")
            self._demo_loop()

    def _demo_loop(self) -> None:
        interval = 1.0 / SAMPLE_RATE
        while self._running:
            t = time.time() - self._t0
            v = (REFERENCE_VOLTAGE
                 + 0.6 * np.sin(2 * np.pi * 0.4 * t)
                 + np.random.normal(0, 0.03))
            with self._lock:
                self._voltage_buf.append(float(v))
                self._time_buf.append(t)
                if self._recording:
                    w = self._cal.to_weight(float(v)) if self._cal.is_valid() else None
                    self._record_buf.append(
                        (t - self._record_start, float(v), w, self._servo_angle))
            time.sleep(interval)

    # ── UI update loop ────────────────────────────────────────────────────────

    def _tick(self) -> None:
        if not self._running:
            return

        with self._lock:
            voltages = list(self._voltage_buf)
            times    = list(self._time_buf)

        if voltages:
            latest = voltages[-1]
            self._v_var.set(f"{latest:+.4f}")

            if self._cal.is_valid():
                self._w_var.set(f"{self._cal.to_weight(latest):.4f}")

            graph_unit = self._graph_unit_var.get()
            if graph_unit == "Voltage" or not self._cal.is_valid():
                plot_data   = voltages
                unit_label  = "Voltage (V)"
                stat_suffix = "V"
            else:
                plot_data   = [self._cal.to_weight_in_unit(v, graph_unit) for v in voltages]
                unit_label  = graph_unit
                stat_suffix = graph_unit

            self._ax.set_ylabel(unit_label, color=TEXT, fontsize=9)
            self._stat_vars["MIN"].set(f"{min(plot_data):.4f} {stat_suffix}")
            self._stat_vars["MAX"].set(f"{max(plot_data):.4f} {stat_suffix}")
            self._stat_vars["AVG"].set(f"{float(np.mean(plot_data)):.4f} {stat_suffix}")

            self._line.set_data(times, plot_data)
            t_end   = times[-1]
            t_start = max(0.0, t_end - HISTORY_SECONDS)
            self._ax.set_xlim(t_start, t_end + 0.2)
            lo, hi = min(plot_data), max(plot_data)
            margin = max(0.15, (hi - lo) * 0.15)
            self._ax.set_ylim(lo - margin, hi + margin)
            self._canvas.draw_idle()

            # Servo angle graph — record current angle at tick rate
            t_now = time.time() - self._t0
            self._servo_angle_buf.append(self._servo_angle)
            self._servo_time_buf.append(t_now)

            angle_times  = list(self._servo_time_buf)
            angle_values = list(self._servo_angle_buf)
            self._angle_line.set_data(angle_times, angle_values)
            self._angle_ax.set_xlim(
                max(0.0, angle_times[-1] - HISTORY_SECONDS), angle_times[-1] + 0.2)
            self._angle_canvas.draw_idle()

        # Keep angle display in sync while waveform is running
        if self._waveform_active:
            self._servo_disp_var.set(f"{self._servo_angle:.1f}°")

        if self._recording:
            self._rec_count_var.set(f"{len(self._record_buf):,} samples")

        self._after_id = self.root.after(100, self._tick)

    # ── Recording ─────────────────────────────────────────────────────────────

    def _toggle_recording(self) -> None:
        if not self._recording:
            if not HAS_XLSX:
                messagebox.showerror("openpyxl not installed",
                                     "Run:  pip install openpyxl\nthen restart the app.")
                return
            default = f"recording_{datetime.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            path = filedialog.asksaveasfilename(
                title="Save recording to…",
                defaultextension=".xlsx",
                initialfile=default,
                filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
            )
            if not path:
                return
            self._record_path = path
            with self._lock:
                self._record_buf.clear()
                self._record_start = time.time() - self._t0
            self._recording = True
            self._rec_btn.configure(text="■ Stop", bg="#592929", fg=TEXT)
            self._rec_count_var.set("0 samples")
        else:
            self._recording = False
            with self._lock:
                buf = list(self._record_buf)
                self._record_buf.clear()
            self._rec_btn.configure(text="● Record", bg=BORDER, fg=TEXT)
            self._rec_count_var.set("")
            self._save_recording(buf)

    def _save_recording(self, buf: list) -> None:
        if not buf:
            messagebox.showinfo("Empty recording", "No samples were captured.")
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Recording"

        hdr_font  = Font(bold=True, color="1E1E2E")
        hdr_fill  = PatternFill("solid", fgColor="89B4FA")
        hdr_align = Alignment(horizontal="center")

        cal_unit = self._cal.unit if self._cal.is_valid() else None
        headers  = (["Time (s)", "Voltage (V)"]
                    + ([f"Weight ({cal_unit})"] if cal_unit else [])
                    + ["Servo Angle (°)"])

        for col, title in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=title)
            cell.font      = hdr_font
            cell.fill      = hdr_fill
            cell.alignment = hdr_align

        for row_idx, row in enumerate(buf, start=2):
            t, v, w, angle = row
            ws.cell(row=row_idx, column=1, value=round(t, 4))
            ws.cell(row=row_idx, column=2, value=round(v, 6))
            if cal_unit and w is not None:
                ws.cell(row=row_idx, column=3, value=round(w, 6))
                ws.cell(row=row_idx, column=4, value=round(angle, 2))
            else:
                ws.cell(row=row_idx, column=3, value=round(angle, 2))

        for col in ws.columns:
            width = max(len(str(cell.value or "")) for cell in col) + 4
            ws.column_dimensions[col[0].column_letter].width = width

        try:
            wb.save(self._record_path)
            messagebox.showinfo(
                "Recording saved",
                f"{len(buf):,} samples saved to:\n{self._record_path}")
        except Exception as exc:
            messagebox.showerror("Save failed", str(exc))

    # ── Helpers ───────────────────────────────────────────────────────────────

    def _clear(self) -> None:
        with self._lock:
            self._voltage_buf.clear()
            self._time_buf.clear()
        self._servo_angle_buf.clear()
        self._servo_time_buf.clear()
        self._t0 = time.time()

    def _on_close(self) -> None:
        if self._recording:
            self._recording = False
            with self._lock:
                buf = list(self._record_buf)
                self._record_buf.clear()
            if buf:
                self._save_recording(buf)
        self._running = False
        if self._after_id is not None:
            self.root.after_cancel(self._after_id)
            self._after_id = None
        self._stop_servo()
        if self._ao_task is not None:
            try:
                self._ao_task.stop()
                self._ao_task.close()
            except Exception:
                pass
        self.root.quit()
        self.root.destroy()


if __name__ == "__main__":
    root = tk.Tk()
    App(root)
    root.mainloop()
